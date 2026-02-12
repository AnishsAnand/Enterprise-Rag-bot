from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File, Request
from pydantic import BaseModel, HttpUrl, Field
from typing import List, Dict, Any, Optional
from app.utils.token_utils import get_user_id_from_request, get_token_from_request
import asyncio
from datetime import datetime
import mimetypes
from io import BytesIO, StringIO
import csv
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
from bs4 import BeautifulSoup
import os
import logging
import re
import urllib.parse
from urllib.parse import urljoin, urlparse
import difflib
import inspect
import json
import hashlib
from typing import Any, Dict, List, Optional, Set, Tuple
from collections import deque
from app.services.scraper_service import scraper_service
from app.services.postgres_service import postgres_service
from app.services.ai_service import ai_service
from app.agents import get_agent_manager  
import numpy as np


router = APIRouter()
logger = logging.getLogger(__name__)

MIN_RELEVANCE_THRESHOLD = float(os.getenv("MIN_RELEVANCE_THRESHOLD", "0.25"))
MAX_CHUNKS_RETURN = int(os.getenv("MAX_CHUNKS_RETURN", "12"))
# ===================== Request Models =====================

class WidgetQueryRequest(BaseModel):
    query: str = Field(..., min_length=1, max_length=500, description="The search query")
    max_results: int = Field(default=50, ge=1, le=100, description="Maximum number of results")
    include_sources: bool = Field(default=True, description="Include source information")
    enable_advanced_search: bool = Field(default=True, description="Enable advanced search features")
    search_depth: str = Field(default="balanced", pattern="^(quick|balanced|deep)$")
    auto_execute: bool = Field(default=True, description="Auto-execute detected tasks")
    store_interaction: bool = Field(default=True, description="Store interaction in knowledge base")
    session_id: Optional[str] = Field(default=None, description="Session ID for multi-turn conversations")
    user_id: Optional[str] = Field(default=None, description="User ID for tracking")
    force_rag_only: bool = Field(default=False, description="Force pure RAG lookup, skip agent routing (used by RAGAgent)")

class WidgetScrapeRequest(BaseModel):
    url: HttpUrl
    store_in_knowledge: bool = True
    extract_images: bool = True
    wait_for_js: bool = False

class BulkScrapeRequest(BaseModel):
    base_url: HttpUrl
    max_depth: int = Field(default=8, ge=1, le=20)
    max_urls: int = Field(default=500, ge=1, le=3000)
    seed_urls: Optional[List[str]] = Field(
        default=None,
        description="Additional URLs to always scrape (merged with discovered URLs)"
    )
    auto_store: bool = True
    domain_filter: Optional[str] = None
    follow_external_links: bool = Field(
        default=False,
        description="Whether the scraper is allowed to follow links outside the base domain"
    )
    extract_deep_images: bool = Field(
        default=False,
        description="Extract images from deeply nested pages"
    )

    extract_tables: bool = Field(
        default=True,
        description="Extract tabular content from pages"
    )

    include_metadata: bool = Field(
        default=True,
        description="Include title, headings, and metadata"
    )

    clean_html: bool = Field(
        default=True,
        description="Strip scripts, styles, nav, and boilerplate"
    )

    chunk_size: int = Field(
        default=512,
        ge=128,
        le=2048,
        description="Chunk size for RAG ingestion"
    )

    chunk_overlap: int = Field(
        default=50,
        ge=0,
        le=500
    )
class EnhancedBulkScraperService:
    """
    Advanced bulk scraping engine with intelligent discovery and parallel processing.
    """

    def __init__(self, scraper_service, max_concurrent: int = 10):
        self.scraper_service = scraper_service
        self.max_concurrent = max_concurrent
        self.semaphore = asyncio.Semaphore(max_concurrent)

        # Crawl state
        self.discovered_urls: Set[str] = set()
        self.visited_urls: Set[str] = set()
        self.failed_urls: Set[str] = set()
        self.content_hashes: Set[str] = set()

        # Priority queue: (priority, depth, url)
        self.url_queue: deque[Tuple[int, int, str]] = deque()

        # Statistics
        self.stats = {
            "total_discovered": 0,
            "total_scraped": 0,
            "total_documents": 0,
            "total_images": 0,
            "start_time": None,
            "end_time": None,
        }
    # ============================================================
    # URL DISCOVERY
    # ============================================================

    async def discover_url(
        self,
        base_url: str,
        max_depth: int = 15,
        max_urls: int = 5000,
        follow_external: bool = False,
        url_patterns: Optional[List[str]] = None,
        exclude_patterns: Optional[List[str]] = None,
    ) -> List[str]:
        """
        Priority-based URL discovery.

        Priority (0–100):
        - 90–100: Docs, guides, APIs
        - 70–89: Product & feature pages
        - 50–69: Blogs & articles
        - <50: Low-value or navigational pages
        """

        self.stats["start_time"] = datetime.now()
        base_domain = urlparse(base_url).netloc

        self.url_queue.append((100, 0, base_url))
        self.discovered_urls.add(base_url)

        logger.info(
            "🔍 Starting intelligent URL discovery | Base=%s Depth=%d URLs=%d",
            base_url,
            max_depth,
            max_urls,
        )

        while self.url_queue and len(self.discovered_urls) < max_urls:
            depth, current_url = self.url_queue.popleft()

            if current_url in self.visited_urls or depth > max_depth:
                continue

            self.visited_urls.add(current_url)

            html = await self._fetch_page_safe(current_url)
            if not html:
                self.failed_urls.add(current_url)
                continue

            soup = BeautifulSoup(html, "html.parser")
            links = self._extract_links_enhanced(soup, current_url)

            for link_url, link_text in links:
                parsed = urlparse(link_url)

                if not follow_external and parsed.netloc != base_domain:
                    continue

                if link_url in self.discovered_urls:
                    continue

                if url_patterns and not self._matches_patterns(link_url, url_patterns):
                    continue

                if exclude_patterns and self._matches_patterns(link_url, exclude_patterns):
                    continue

                score = self._calculate_url_priority(link_url, link_text, depth + 1)
                self.url_queue.append((score, depth + 1, link_url))
                self.discovered_urls.add(link_url)

            self.url_queue = deque(
                sorted(self.url_queue, key=lambda x: x[0], reverse=True)
            )

            if len(self.discovered_urls) % 100 == 0:
                logger.info(
                    "📊 Discovered=%d Visited=%d Queue=%d",
                    len(self.discovered_urls),
                    len(self.visited_urls),
                    len(self.url_queue),
                )

            await asyncio.sleep(0.1)

        self.stats["total_discovered"] = len(self.discovered_urls)
        return list(self.discovered_urls)

    # ============================================================
    # PARALLEL SCRAPING
    # ============================================================

    async def bulk_scrape_parallel(
        self,
        urls: List[str],
        extract_images: bool = True,
        extract_tables: bool = True,
        chunk_size: int = 2000,
        chunk_overlap: int = 300,
    ) -> List[Dict[str, Any]]:
        """
        Parallel bulk scraping with quality filtering and intelligent chunking.
        """

        documents: List[Dict[str, Any]] = []
        batch_size = 50

        for i in range(0, len(urls), batch_size):
            batch = urls[i : i + batch_size]

            tasks = [
                self._scrape_url_enhanced(url, extract_images, extract_tables)
                for url in batch
            ]

            results = await asyncio.gather(*tasks, return_exceptions=True)

            for url, result in zip(batch, results):
                if isinstance(result, Exception) or not result:
                    continue

                text = result["content"].get("text", "")
                images = result["content"].get("images", [])

                if not self._is_quality_content(text, url):
                    continue

                fingerprint = self._simhash(text)
                if fingerprint in self.content_hashes:
                    continue

                self.content_hashes.add(fingerprint)
                chunks = self._chunk_text_intelligent(text, chunk_size, chunk_overlap)

                for idx, chunk in enumerate(chunks):
                    documents.append(
                        {
                            "content": chunk,
                            "url": f"{url}#chunk-{idx}" if idx else url,
                            "title": result["content"].get("title", ""),
                            "timestamp": datetime.now().isoformat(),
                            "images": images if idx == 0 else [],
                            "metadata": {
                                "chunk_index": idx,
                                "total_chunks": len(chunks),
                                "quality": self._calculate_content_quality(text),
                                "word_count": len(chunk.split()),
                            },
                        }
                    )

                self.stats["total_scraped"] += 1
                self.stats["total_documents"] += len(chunks)
                self.stats["total_images"] += len(images)

            await asyncio.sleep(1)

        self.stats["end_time"] = datetime.now()
        return documents

    # ============================================================
    # HELPERS
    # ============================================================

    async def _fetch_page_safe(self, url: str) -> Optional[str]:
        for attempt in range(3):
            try:
                async with self.semaphore:
                    r = self.scraper_service.session.get(
                        url, timeout=self.scraper_service.request_timeout
                    )
                    r.raise_for_status()
                    if "text/html" in r.headers.get("Content-Type", ""):
                        return r.text
            except Exception:
                await asyncio.sleep(2**attempt)
        return None

    def _extract_links_enhanced(
        self, soup: BeautifulSoup, base_url: str
    ) -> List[Tuple[str, str]]:
        links = []
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if href.startswith(("#", "javascript:", "mailto:")):
                continue

            url = urljoin(base_url, href)
            parsed = urlparse(url)._replace(fragment="")
            clean = parsed.geturl()

            if clean.lower().endswith(
                (".pdf", ".jpg", ".png", ".css", ".js", ".zip", ".svg")
            ):
                continue

            links.append((clean, a.get_text(strip=True)))
        return links

    def _calculate_url_priority(self, url: str, text: str, depth: int) -> int:
        score = 50
        high = ["doc", "guide", "tutorial", "api", "sdk"]
        medium = ["feature", "product", "integration"]
        low = ["tag", "category", "login", "search"]

        if any(k in url.lower() or k in text.lower() for k in high):
            score += 40
        elif any(k in url.lower() or k in text.lower() for k in medium):
            score += 20

        if any(k in url.lower() for k in low):
            score -= 30

        score -= depth * 5
        return max(0, min(100, score))

    def _matches_patterns(self, url: str, patterns: List[str]) -> bool:
        return any(re.search(p, url, re.IGNORECASE) for p in patterns)

    def _simhash(self, text: str) -> str:
        words = re.findall(r"\w+", text.lower())
        return hashlib.md5(" ".join(words[:1000]).encode()).hexdigest()[:16]

    def _is_quality_content(self, text: str, _: str) -> bool:
        return len(text.split()) >= 50

    def _chunk_text_intelligent(
        self, text: str, size: int, overlap: int
    ) -> List[str]:
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        chunks, current = [], ""

        for p in paragraphs:
            if len(current) + len(p) > size:
                chunks.append(current.strip())
                current = current[-overlap:] + "\n\n" + p
            else:
                current += "\n\n" + p if current else p

        if current.strip():
            chunks.append(current.strip())
        return chunks

    def _calculate_content_quality(self, text: str) -> float:
        wc = len(text.split())
        diversity = len(set(text.lower().split())) / max(wc, 1)
        return min(1.0, 0.4 + diversity)

class TaskExecutionRequest(BaseModel):
    task_description: str = Field(..., min_length=1, max_length=2000)
    context: Optional[Dict[str, Any]] = None
    auto_connect_services: bool = Field(default=True)
    store_result: bool = Field(default=True)


# ===================== Scraping Service =====================
from app.services.scraper_service import scraper_service
class ProductionBulkScraperService:
    """
    ✅ PRODUCTION-READY: Enhanced bulk scraper with intelligent discovery.
    
    NEW FEATURES:
    - Priority-based URL discovery (docs > products > blogs)
    - Parallel scraping with adaptive batch sizing
    - Content quality filtering
    - Duplicate detection via simhash
    - Comprehensive error handling
    """

    def __init__(self, scraper_service, max_concurrent: int = 10):
        self.scraper_service = scraper_service
        self.max_concurrent = max_concurrent
        self.semaphore = asyncio.Semaphore(max_concurrent)

        # Crawl state
        self.discovered_urls: Set[str] = set()
        self.visited_urls: Set[str] = set()
        self.failed_urls: Set[str] = set()
        self.content_hashes: Set[str] = set()

        # Priority queue: (priority, depth, url)
        self.url_queue: deque = deque()

        # Statistics
        self.stats = {
            "total_discovered": 0,
            "total_scraped": 0,
            "total_documents": 0,
            "total_images": 0,
            "start_time": None,
            "end_time": None,
        }

    async def discover_url(
        self,
        base_url: str,
        max_depth: int = 15,
        max_urls: int = 5000,
        follow_external: bool = False,
        url_patterns: Optional[List[str]] = None,
        exclude_patterns: Optional[List[str]] = None,
    ) -> List[str]:
        """
        ✅ Priority-based URL discovery with intelligent scoring.
        
        Priority Tiers:
        - 90-100: Documentation, guides, APIs, SDKs
        - 70-89: Product pages, features, integrations
        - 50-69: Blog posts, articles, news
        - <50: Navigation pages, tags, categories
        """
        self.stats["start_time"] = datetime.now()
        base_domain = urlparse(base_url).netloc

        # Initialize queue with base URL
        self.url_queue.append((100, 0, base_url))
        self.discovered_urls.add(base_url)

        logger.info(
            f"🔍 Starting intelligent URL discovery\n"
            f"   Base: {base_url}\n"
            f"   Max depth: {max_depth}\n"
            f"   Max URLs: {max_urls}\n"
            f"   Follow external: {follow_external}"
        )

        while self.url_queue and len(self.discovered_urls) < max_urls:
            # Get highest priority URL
            priority, depth, current_url = self.url_queue.popleft()

            # Skip if already visited or too deep
            if current_url in self.visited_urls or depth > max_depth:
                continue

            self.visited_urls.add(current_url)

            # Fetch page with retry logic
            html = await self._fetch_page_safe(current_url)
            if not html:
                self.failed_urls.add(current_url)
                continue

            # Extract links
            soup = BeautifulSoup(html, "html.parser")
            links = self._extract_links_enhanced(soup, current_url)

            # Process each discovered link
            for link_url, link_text in links:
                parsed = urlparse(link_url)

                # Domain filtering
                if not follow_external and parsed.netloc != base_domain:
                    continue

                # Skip if already discovered
                if link_url in self.discovered_urls:
                    continue

                # Pattern filtering
                if url_patterns and not self._matches_patterns(link_url, url_patterns):
                    continue

                if exclude_patterns and self._matches_patterns(link_url, exclude_patterns):
                    continue

                # Calculate priority score
                score = self._calculate_url_priority(link_url, link_text, depth + 1)
                
                # Add to queue
                self.url_queue.append((score, depth + 1, link_url))
                self.discovered_urls.add(link_url)

            # Re-sort queue by priority (keep highest priority first)
            self.url_queue = deque(
                sorted(self.url_queue, key=lambda x: x[0], reverse=True)
            )

            # Progress logging
            if len(self.discovered_urls) % 100 == 0:
                logger.info(
                    f"📊 Discovery progress: {len(self.discovered_urls)} URLs found, "
                    f"{len(self.visited_urls)} visited, "
                    f"{len(self.url_queue)} queued"
                )

            # Small delay to avoid overwhelming the server
            await asyncio.sleep(0.1)

        self.stats["total_discovered"] = len(self.discovered_urls)
        
        logger.info(
            f"✅ URL discovery complete\n"
            f"   Total discovered: {len(self.discovered_urls)}\n"
            f"   Total visited: {len(self.visited_urls)}\n"
            f"   Failed: {len(self.failed_urls)}"
        )

        return list(self.discovered_urls)

    async def bulk_scrape_parallel(
        self,
        urls: List[str],
        extract_images: bool = True,
        extract_tables: bool = True,
        chunk_size: int = 2000,
        chunk_overlap: int = 300,
    ) -> List[Dict[str, Any]]:
        """
        ✅ Parallel bulk scraping with quality filtering and intelligent chunking.
        """
        documents: List[Dict[str, Any]] = []
        batch_size = 50

        logger.info(
            f"🚀 Starting parallel bulk scrape\n"
            f"   Total URLs: {len(urls)}\n"
            f"   Batch size: {batch_size}\n"
            f"   Extract images: {extract_images}"
        )

        for i in range(0, len(urls), batch_size):
            batch = urls[i : i + batch_size]
            
            logger.info(
                f"📦 Processing batch {i//batch_size + 1}/{(len(urls)-1)//batch_size + 1} "
                f"({len(batch)} URLs)"
            )

            # Create scraping tasks
            tasks = [
                self._scrape_url_enhanced(url, extract_images, extract_tables)
                for url in batch
            ]

            # Execute batch with exception handling
            results = await asyncio.gather(*tasks, return_exceptions=True)

            # Process results
            for url, result in zip(batch, results):
                # Skip exceptions
                if isinstance(result, Exception):
                    logger.warning(f"⚠️ Scraping failed for {url}: {result}")
                    continue

                if not result:
                    continue

                # Extract content
                text = result["content"].get("text", "")
                images = result["content"].get("images", [])

                # Quality filtering
                if not self._is_quality_content(text, url):
                    logger.debug(f"Skipping low-quality content: {url[:60]}")
                    continue

                # Duplicate detection via content hashing
                fingerprint = self._simhash(text)
                if fingerprint in self.content_hashes:
                    logger.debug(f"Skipping duplicate content: {url[:60]}")
                    continue

                self.content_hashes.add(fingerprint)

                # Intelligent chunking
                chunks = self._chunk_text_intelligent(text, chunk_size, chunk_overlap)

                # Create document entries
                for idx, chunk in enumerate(chunks):
                    documents.append({
                        "content": chunk,
                        "url": f"{url}#chunk-{idx}" if idx else url,
                        "title": result["content"].get("title", ""),
                        "timestamp": datetime.now().isoformat(),
                        "images": images if idx == 0 else [],  # Only first chunk gets images
                        "metadata": {
                            "chunk_index": idx,
                            "total_chunks": len(chunks),
                            "quality": self._calculate_content_quality(text),
                            "word_count": len(chunk.split()),
                        },
                    })

                self.stats["total_scraped"] += 1
                self.stats["total_documents"] += len(chunks)
                self.stats["total_images"] += len(images)

            # Delay between batches
            await asyncio.sleep(1)

        self.stats["end_time"] = datetime.now()
        
        logger.info(
            f"✅ Bulk scrape complete\n"
            f"   Scraped: {self.stats['total_scraped']} pages\n"
            f"   Documents: {self.stats['total_documents']}\n"
            f"   Images: {self.stats['total_images']}"
        )

        return documents

    # ========================================================================
    # HELPER METHODS
    # ========================================================================

    async def _fetch_page_safe(self, url: str) -> Optional[str]:
        """Fetch page with retry logic and error handling."""
        for attempt in range(3):
            try:
                async with self.semaphore:
                    r = self.scraper_service.session.get(
                        url, timeout=self.scraper_service.request_timeout
                    )
                    r.raise_for_status()
                    if "text/html" in r.headers.get("Content-Type", ""):
                        return r.text
            except Exception as e:
                if attempt == 2:
                    logger.debug(f"Failed to fetch {url} after 3 attempts: {e}")
                await asyncio.sleep(2**attempt)
        return None

    def _extract_links_enhanced(
        self, soup: BeautifulSoup, base_url: str
    ) -> List[tuple]:
        """Extract and normalize links from HTML."""
        links = []
        
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            
            # Skip invalid links
            if href.startswith(("#", "javascript:", "mailto:", "tel:")):
                continue

            # Resolve relative URLs
            url = urljoin(base_url, href)
            parsed = urlparse(url)._replace(fragment="")
            clean = parsed.geturl()

            # Skip file downloads
            if clean.lower().endswith((
                ".pdf", ".jpg", ".png", ".gif", ".css", ".js", 
                ".zip", ".tar", ".gz", ".svg", ".ico"
            )):
                continue

            # Get link text for priority calculation
            link_text = a.get_text(strip=True)
            
            links.append((clean, link_text))
        
        return links

    def _calculate_url_priority(self, url: str, text: str, depth: int) -> int:
        """
        Calculate URL priority score (0-100).
        
        Priority factors:
        - High-value keywords (docs, api, guide) → +40
        - Medium-value keywords (product, feature) → +20
        - Low-value keywords (tag, category) → -30
        - Depth penalty → -5 per level
        """
        score = 50  # Base score
        
        # High-value keywords
        high_value = ["doc", "guide", "tutorial", "api", "sdk", "reference", "manual"]
        medium_value = ["feature", "product", "integration", "solution"]
        low_value = ["tag", "category", "login", "search", "archive"]

        url_lower = url.lower()
        text_lower = text.lower()

        # Check keywords in URL and text
        if any(k in url_lower or k in text_lower for k in high_value):
            score += 40
        elif any(k in url_lower or k in text_lower for k in medium_value):
            score += 20

        if any(k in url_lower for k in low_value):
            score -= 30

        # Depth penalty (prioritize shallower pages)
        score -= depth * 5

        return max(0, min(100, score))

    def _matches_patterns(self, url: str, patterns: List[str]) -> bool:
        """Check if URL matches any of the provided regex patterns."""
        return any(re.search(p, url, re.IGNORECASE) for p in patterns)

    def _simhash(self, text: str) -> str:
        """Generate content fingerprint for duplicate detection."""
        words = re.findall(r"\w+", text.lower())
        return hashlib.md5(" ".join(words[:1000]).encode()).hexdigest()[:16]

    def _is_quality_content(self, text: str, url: str) -> bool:
        """Check if content meets quality threshold."""
        # Minimum word count
        if len(text.split()) < 50:
            return False
        
        # Check for boilerplate content
        boilerplate_indicators = [
            "403 forbidden",
            "404 not found",
            "access denied",
            "javascript required",
            "cookies required"
        ]
        
        text_lower = text.lower()
        if any(indicator in text_lower for indicator in boilerplate_indicators):
            return False
        
        return True

    def _chunk_text_intelligent(
        self, text: str, size: int, overlap: int
    ) -> List[str]:
        """
        Intelligent text chunking that respects paragraph boundaries.
        """
        # Split into paragraphs
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        
        chunks = []
        current = ""

        for p in paragraphs:
            # If adding this paragraph exceeds size, start new chunk
            if len(current) + len(p) > size and current:
                chunks.append(current.strip())
                # Keep overlap from end of previous chunk
                current = current[-overlap:] + "\n\n" + p
            else:
                current += "\n\n" + p if current else p

        # Add final chunk
        if current.strip():
            chunks.append(current.strip())
        
        return chunks if chunks else [text]

    def _calculate_content_quality(self, text: str) -> float:
        """Calculate content quality score (0-1)."""
        word_count = len(text.split())
        unique_words = len(set(text.lower().split()))
        
        # Lexical diversity
        diversity = unique_words / max(word_count, 1)
        
        # Base quality score
        quality = min(1.0, 0.4 + diversity)
        
        return round(quality, 2)

    async def _scrape_url_enhanced(
        self,
        url: str,
        extract_images: bool,
        extract_tables: bool
    ) -> Optional[Dict]:
        """
        Enhanced single URL scraping with comprehensive extraction.
        """
        try:
            scrape_params = {
                "extract_text": True,
                "extract_links": True,
                "extract_images": extract_images,
                "extract_tables": extract_tables,
                "scroll_page": True,
                "wait_for_element": "body",
                "output_format": "json",
            }
            
            result = await call_maybe_async(
                self.scraper_service.scrape_url,
                url,
                scrape_params
            )
            
            return result
        
        except Exception as e:
            logger.debug(f"Scraping failed for {url}: {e}")
            return None

# Global instance
enhanced_bulk_scraper = ProductionBulkScraperService(scraper_service)

async def call_maybe_async(fn, *args, **kwargs):
    """Call a function that might be sync or async."""
    import inspect
    
    if not callable(fn):
        raise RuntimeError("Provided object is not callable")
    
    result = fn(*args, **kwargs)
    
    if inspect.isawaitable(result):
        return await result
    
    return result

# ===================== RAG SEARCH Service =====================

class EnhancedRAGSearchService:
    

    def __init__(self, postgres_service, ai_service):
        """Initialize with service dependencies."""
        self.postgres_service = postgres_service
        self.ai_service = ai_service
        
        # Base configuration
        self.base_threshold = 0.35  # Base relevance threshold
        self.enable_query_expansion = True
        self.enable_reranking = True
        self.max_chunks_return = 10
        
        logger.info("✅ EnhancedRAGSearchService initialized")
        logger.info(f"   - Base threshold: {self.base_threshold}")
        logger.info(f"   - Query expansion: {self.enable_query_expansion}")
        logger.info(f"   - Semantic reranking: {self.enable_reranking}")

    def _analyze_query_complexity(self, query: str) -> Dict:
        """
        ✅ Analyze query complexity for adaptive thresholding.
        
        Returns:
            {
                "word_count": int,
                "has_technical_terms": bool,
                "has_specific_entities": bool,
                "question_type": "specific" | "general" | "vague",
                "specificity_score": float  # 0.0-1.0
            }
        """
        query_lower = query.lower()
        words = query.split()
        
        # Detect technical terms
        has_acronyms = bool(re.search(r'\b[A-Z]{2,}\b', query))
        has_versions = bool(re.search(r'\d+\.\d+(?:\.\d+)?', query))
        has_technical_terms = has_acronyms or has_versions
        
        # Detect specific entities
        has_proper_nouns = bool(re.search(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\b', query))
        has_specific_numbers = bool(re.search(r'\b\d+\b', query))
        has_specific_entities = has_proper_nouns or has_specific_numbers or has_versions
        
        # Determine question type
        if any(word in query_lower for word in ['what is', 'define', 'explain', 'how does']):
            if len(words) > 8 or has_specific_entities:
                question_type = "specific"
            else:
                question_type = "general"
        elif any(word in query_lower for word in ['how to', 'steps', 'guide', 'tutorial']):
            question_type = "specific"
        elif len(words) <= 3 and not has_specific_entities:
            question_type = "vague"
        elif has_specific_entities or has_technical_terms:
            question_type = "specific"
        else:
            question_type = "general"
        
        # Calculate specificity score
        specificity_score = 0.0
        
        if len(words) > 10:
            specificity_score += 0.3
        elif len(words) > 5:
            specificity_score += 0.15
        
        if has_specific_entities:
            specificity_score += 0.3
        
        if has_technical_terms:
            specificity_score += 0.2
        
        if question_type == "specific":
            specificity_score += 0.2
        elif question_type == "general":
            specificity_score += 0.1
        
        specificity_score = min(1.0, specificity_score)
        
        return {
            "word_count": len(words),
            "has_technical_terms": has_technical_terms,
            "has_specific_entities": has_specific_entities,
            "question_type": question_type,
            "specificity_score": specificity_score
        }

    def _calculate_adaptive_threshold(self, query_complexity: Dict) -> float:
        """
        ✅ Calculate adaptive relevance threshold.
        
        Strategy:
        - Specific queries → HIGH threshold (0.50-0.75)
        - General queries → MEDIUM threshold (0.35-0.50)
        - Vague queries → LOWER threshold (0.30-0.40)
        
        Returns:
            float: Adaptive threshold between 0.30 and 0.75
        """
        base = self.base_threshold  # 0.35
        
        # Adjust based on specificity
        if query_complexity["has_specific_entities"]:
            base += 0.15
        
        if query_complexity["has_technical_terms"]:
            base += 0.10
        
        if query_complexity["question_type"] == "specific":
            base += 0.10
        elif query_complexity["question_type"] == "vague":
            base -= 0.05
        
        # Apply specificity score
        specificity_bonus = query_complexity["specificity_score"] * 0.15
        base += specificity_bonus
        
        # Clamp to range
        adaptive_threshold = max(0.30, min(0.75, base))
        
        logger.info(
            f"[Adaptive Threshold] {adaptive_threshold:.2f} "
            f"(base: {self.base_threshold}, "
            f"specificity: {query_complexity['specificity_score']:.2f}, "
            f"type: {query_complexity['question_type']})"
        )
        
        return adaptive_threshold

    def _cross_reference_validate(
        self, 
        query: str, 
        chunks: List[Dict]
    ) -> List[Dict]:
        """
        ✅ Cross-reference validation for accuracy.
        
        Strategy:
        - Extract key terms from query
        - Check term coverage in each chunk
        - Count peer support (how many chunks agree)
        - Boost confidence for supported claims
        - Penalize isolated claims
        """
        if len(chunks) <= 1:
            return chunks
        
        # Extract query terms (filter stopwords)
        stopwords = {
            'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
            'of', 'with', 'by', 'from', 'as', 'is', 'was', 'are', 'were'
        }
        
        query_terms = [
            word for word in re.findall(r'\b\w+\b', query.lower())
            if word not in stopwords and len(word) > 2
        ]
        
        validated_chunks = []
        
        for i, chunk in enumerate(chunks):
            chunk_text = chunk.get("content", "").lower()
            chunk_copy = chunk.copy()
            
            # Calculate term coverage
            terms_found = sum(1 for term in query_terms if term in chunk_text)
            term_coverage = terms_found / len(query_terms) if query_terms else 0
            
            # Calculate peer support
            peer_support_count = 0
            for j, other_chunk in enumerate(chunks):
                if i == j:
                    continue
                
                other_text = other_chunk.get("content", "").lower()
                
                # Check for shared key terms
                shared_terms = sum(
                    1 for term in query_terms 
                    if term in chunk_text and term in other_text
                )
                
                if shared_terms >= max(1, len(query_terms) * 0.3):
                    peer_support_count += 1
            
            # Adjust confidence
            original_confidence = chunk.get("relevance_score", 0.5)
            
            # Boost for high term coverage
            if term_coverage >= 0.7:
                original_confidence *= 1.1
            elif term_coverage >= 0.4:
                original_confidence *= 1.05
            
            # Boost for peer support
            if peer_support_count >= 2:
                original_confidence *= 1.1
            elif peer_support_count >= 1:
                original_confidence *= 1.05
            
            # Penalize for low coverage and no peer support
            if term_coverage < 0.2 and peer_support_count == 0:
                original_confidence *= 0.9
            
            chunk_copy["relevance_score"] = min(1.0, original_confidence)
            chunk_copy["validation_metadata"] = {
                "term_coverage": term_coverage,
                "peer_support_count": peer_support_count,
                "terms_found": terms_found,
                "total_terms": len(query_terms)
            }
            
            validated_chunks.append(chunk_copy)
        
        logger.info(
            f"[Cross-Reference] Validated {len(validated_chunks)} chunks "
            f"(avg peer support: {np.mean([c.get('validation_metadata', {}).get('peer_support_count', 0) for c in validated_chunks]):.1f})"
        )
        
        return validated_chunks
    
    async def search(
        self,
        query: str,
        top_k: int = 10,
        user_id: Optional[int] = None
    ) -> Dict:
        """
        ✅ Comprehensive RAG search with adaptive intelligence.
        
        Flow:
        1. Analyze query complexity
        2. Calculate adaptive threshold
        3. Search PostgreSQL (hybrid vector + FTS)
        4. Cross-reference validate results
        5. Apply semantic reranking
        6. Filter by adaptive threshold (two-pass)
        7. Return top-k results
        
        Returns:
            {
                "chunks": [...],
                "total_results": int,
                "search_quality": "high" | "medium" | "low",
                "query_complexity": {...},
                "adaptive_threshold": float,
                "execution_time_ms": float
            }
        """
        start_time = datetime.utcnow()
        
        # Step 1: Analyze query complexity
        query_complexity = self._analyze_query_complexity(query)
        adaptive_threshold = self._calculate_adaptive_threshold(query_complexity)
        
        logger.info(
            f"[RAG Search] Query: '{query[:100]}'\n"
            f"   Complexity: {query_complexity['question_type']} "
            f"(specificity: {query_complexity['specificity_score']:.2f})\n"
            f"   Adaptive threshold: {adaptive_threshold:.2f}"
        )
        
        try:
            # Step 2: Query expansion (optional)
            if self.enable_query_expansion:
                expanded_query = await self._expand_query(query)
            else:
                expanded_query = query
            
            # Step 3: Search PostgreSQL
            initial_results = await self.postgres_service.search_documents(
                query=expanded_query,
                n_results=top_k * 5  # Get more for filtering
            )
            
            if not initial_results:
                logger.warning(f"[RAG Search] No results found for query: '{query}'")
                return self._create_no_results_response(query, start_time)
            
            logger.info(f"[RAG Search] Retrieved {len(initial_results)} initial results")
            
            # Step 4: Cross-reference validation
            validated_results = self._cross_reference_validate(query, initial_results)
            
            # Step 5: Semantic reranking (if enabled)
            if self.enable_reranking:
                reranked_results = await self._semantic_rerank(query, validated_results)
            else:
                reranked_results = validated_results
            
            # Step 6: Two-pass threshold filtering
            # Pass 1: Try strict threshold
            strict_filtered = [
                chunk for chunk in reranked_results 
                if chunk.get("relevance_score", 0) >= adaptive_threshold
            ]
            
            # Pass 2: If too few results, relax threshold
            if len(strict_filtered) < 3:
                relaxed_threshold = max(0.30, adaptive_threshold - 0.15)
                final_results = [
                    chunk for chunk in reranked_results 
                    if chunk.get("relevance_score", 0) >= relaxed_threshold
                ][:top_k]
                logger.info(
                    f"[RAG Search] Relaxed threshold to {relaxed_threshold:.2f} "
                    f"(found {len(final_results)} chunks)"
                )
            else:
                final_results = strict_filtered[:top_k]
                logger.info(
                    f"[RAG Search] Strict threshold {adaptive_threshold:.2f} "
                    f"yielded {len(final_results)} chunks"
                )
            
            execution_time = (datetime.utcnow() - start_time).total_seconds() * 1000
            
            logger.info(
                f"[RAG Search] ✅ Returning {len(final_results)} results "
                f"in {execution_time:.1f}ms\n"
                f"   Quality: {self._assess_quality(final_results)}\n"
                f"   Avg confidence: {np.mean([c.get('relevance_score', 0) for c in final_results]):.2f}"
            )
            
            return {
                "chunks": final_results,
                "total_results": len(final_results),
                "search_quality": self._assess_quality(final_results),
                "query_complexity": query_complexity,
                "adaptive_threshold": adaptive_threshold,
                "execution_time_ms": execution_time,
                "metadata": {
                    "initial_results": len(initial_results),
                    "after_validation": len(validated_results),
                    "after_reranking": len(reranked_results),
                    "after_threshold": len(final_results),
                }
            }
        
        except Exception as e:
            logger.exception(f"[RAG Search] Error: {e}")
            raise

    async def _expand_query(self, query: str) -> str:
        """Expand query with related terms."""
        try:
            prompt = f"""Generate 3 alternative search terms for this query (return only terms, comma-separated):
Query: "{query}"

Terms:"""
            
            response = await self.ai_service.generate_response(prompt, [])
            
            # Parse response
            if isinstance(response, str):
                expanded_terms = [t.strip() for t in response.split(",")]
                return f"{query} {' '.join(expanded_terms[:3])}"
            
            return query
        
        except Exception:
            return query

    async def _semantic_rerank(self,query: str,results: List[Dict]) -> List[Dict]:
        """Apply semantic reranking using stored embeddings (NO re-embedding)."""
        if query in self.query_embedding_cache:
            query_embedding = self.query_embedding_cache[query]
        else:
            emb = await self.ai_service.generate_embeddings([query])
            query_embedding = np.array(emb[0])
            self.query_embedding_cache[query] = query_embedding

        if not results:
            return []

        try:
        # ✅ Generate query embedding ONCE
            query_embeddings = await self.ai_service.generate_embeddings([query])
            if not query_embeddings:
                return results

            query_embedding = np.array(query_embeddings[0])
            query_norm = np.linalg.norm(query_embedding) + 1e-10

            reranked = []

            for result in results:
                doc_embedding = result.get("embedding")
                if not doc_embedding:
                    reranked.append(result)
                    continue

                doc_embedding = np.array(doc_embedding)
                doc_norm = np.linalg.norm(doc_embedding) + 1e-10

            # ✅ Cosine similarity (local, fast)
                similarity = float(
                np.dot(query_embedding, doc_embedding) / (query_norm * doc_norm)
            )

                original_score = result.get("relevance_score", 0.5)
                final_score = 0.6 * similarity + 0.4 * original_score

                r = result.copy()
                r["semantic_score"] = similarity
                r["relevance_score"] = final_score
                reranked.append(r)

            reranked.sort(key=lambda x: x["relevance_score"], reverse=True)
            return reranked

        except Exception as e:
            logger.error(f"Reranking error: {e}")
            return results

    def _assess_quality(self, chunks: List[Dict]) -> str:
        """Assess search quality."""
        if not chunks:
            return "low"
        
        avg_score = np.mean([c.get("relevance_score", 0) for c in chunks])
        
        if avg_score >= 0.8:
            return "high"
        elif avg_score >= 0.6:
            return "medium"
        else:
            return "low"

    def _create_no_results_response(self, query: str, start_time: datetime) -> Dict:
        """Create response for no results."""
        execution_time = (datetime.utcnow() - start_time).total_seconds() * 1000
        
        return {
            "chunks": [],
            "total_results": 0,
            "search_quality": "low",
            "no_results": True,
            "no_results_message": (
                f"No relevant information found for '{query}' in knowledge base."
            ),
            "execution_time_ms": execution_time,
            "metadata": {}
        }
    
# ===================== Orchestration Service =====================

class OrchestrationService:
    """
    Advanced orchestration service for automatic task detection,
    service connection, and intelligent task execution.
    """
    
    def __init__(self):
        self.task_patterns = {
            "scrape": [
                r"scrape\s+(?:the\s+)?(?:website|url|page)\s+(.+)",
                r"extract\s+(?:data|content|information)\s+from\s+(.+)",
                r"fetch\s+(?:data|content)\s+from\s+(.+)",
                r"get\s+information\s+from\s+(.+)",
            ],
            "search": [
                r"search\s+(?:for|about)\s+(.+)",
                r"find\s+(?:information|data)\s+(?:about|on)\s+(.+)",
                r"lookup\s+(.+)",
                r"what\s+(?:is|are)\s+(.+)",
                r"tell\s+me\s+about\s+(.+)",
            ],
            "analyze": [
                r"analyze\s+(.+)",
                r"summarize\s+(.+)",
                r"explain\s+(.+)",
                r"compare\s+(.+)",
            ],
            "upload": [
                r"process\s+(?:the\s+)?(?:file|document)\s+(.+)",
                r"upload\s+(.+)",
                r"read\s+(?:the\s+)?(?:file|document)\s+(.+)",
            ],
            "bulk_operation": [
                r"scrape\s+(?:all|multiple)\s+(?:pages|urls)\s+from\s+(.+)",
                r"bulk\s+(?:scrape|extract)\s+from\s+(.+)",
                r"crawl\s+(.+)",
            ]
        }
        
        self.service_availability = {}
        self.task_history = []
        
    async def detect_task_intent(self, query: str) -> Dict[str, Any]:
        """
        Detect user intent and extract task parameters from natural language.
        """
        query_lower = query.lower().strip()
        
        detected_tasks = []
        
        for task_type, patterns in self.task_patterns.items():
            for pattern in patterns:
                match = re.search(pattern, query_lower, re.IGNORECASE)
                if match:
                    detected_tasks.append({
                        "type": task_type,
                        "confidence": 0.8 if len(match.groups()) > 0 else 0.6,
                        "extracted_params": match.groups(),
                        "original_query": query
                    })
        
        # Use AI to enhance intent detection if no clear pattern match
        if not detected_tasks:
            try:
                ai_intent = await ai_service.detect_task_intent(query)
                if ai_intent and ai_intent.get("confidence", 0) > 0.5:
                    detected_tasks.append(ai_intent)
            except Exception as e:
                logger.warning(f"AI intent detection failed: {e}")
        
        return {
            "query": query,
            "detected_tasks": detected_tasks,
            "primary_task": detected_tasks[0] if detected_tasks else None,
            "requires_clarification": len(detected_tasks) == 0 or (
                len(detected_tasks) > 1 and 
                abs(detected_tasks[0]["confidence"] - detected_tasks[1]["confidence"]) < 0.2
            )
        }
    
    async def check_service_availability(self) -> Dict[str, Dict[str, Any]]:
        """
        Check availability of all integrated services.
        """
        services = {
            "postgres": {"available": False, "status": "unknown", "error": None},
            "ai_service": {"available": False, "status": "unknown", "error": None},
            "scraper": {"available": False, "status": "unknown", "error": None},
        }
        
        # Check postgres
        try:
            stats = await call_maybe_async(postgres_service.get_collection_stats)
            if isinstance(stats, dict) and stats.get("status") in ("active", "healthy"):
                services["postgres"]["available"] = True
                services["postgres"]["status"] = "healthy"
                services["postgres"]["documents"] = stats.get("document_count", 0)
            else:
                services["postgres"]["status"] = "degraded"
        except Exception as e:
            services["postgres"]["error"] = str(e)
            logger.warning(f"postgres availability check failed: {e}")
        
        # Check AI Service
        try:
            health = await ai_service.get_service_health()
            if isinstance(health, dict):
                is_healthy = health.get("overall_status") == "healthy"
                services["ai_service"]["available"] = is_healthy
                services["ai_service"]["status"] = health.get("overall_status", "unknown")
                services["ai_service"]["model"] = health.get("service", {}).get("current_model")
        except Exception as e:
            services["ai_service"]["error"] = str(e)
            logger.warning(f"AI service availability check failed: {e}")
        
        # Check Scraper
        try:
            test_result = await call_maybe_async(
                scraper_service.scrape_url,
                "https://example.com",
                {"extract_text": True, "output_format": "json"}
            )
            services["scraper"]["available"] = test_result is not None
            services["scraper"]["status"] = "healthy" if test_result else "unknown"
        except Exception as e:
            services["scraper"]["error"] = str(e)
            logger.debug(f"Scraper availability check failed (expected): {e}")
        
        self.service_availability = services
        return services
    
    async def execute_orchestrated_task(
        self,
        task_info: Dict[str, Any],
        background_tasks: BackgroundTasks
    ) -> Dict[str, Any]:
        """
        Execute detected task by orchestrating multiple services.
        """
        task_type = task_info.get("type")
        params = task_info.get("extracted_params", [])
        original_query = task_info.get("original_query", "")
        
        execution_log = {
            "task_type": task_type,
            "started_at": datetime.now().isoformat(),
            "steps": [],
            "result": None,
            "error": None
        }
        
        try:
            # Check service availability
            services = await self.check_service_availability()
            execution_log["steps"].append({
                "step": "service_availability_check",
                "timestamp": datetime.now().isoformat(),
                "services": services
            })
            
            # Route to appropriate handler
            if task_type == "scrape":
                result = await self._handle_scrape_task(params, services, background_tasks)
            elif task_type == "search":
                result = await self._handle_search_task(params, services, original_query)
            elif task_type == "analyze":
                result = await self._handle_analyze_task(params, services, original_query)
            elif task_type == "bulk_operation":
                result = await self._handle_bulk_task(params, services, background_tasks)
            else:
                result = await self._handle_generic_task(original_query, services)
            
            execution_log["result"] = result
            execution_log["completed_at"] = datetime.now().isoformat()
            execution_log["status"] = "success"
            
        except Exception as e:
            logger.exception(f"Task execution failed: {e}")
            execution_log["error"] = str(e)
            execution_log["status"] = "failed"
            execution_log["completed_at"] = datetime.now().isoformat()
        
        # Store execution log in history
        self.task_history.append(execution_log)
        
        return execution_log
    
    async def _handle_scrape_task(
        self,
        params: List[str],
        services: Dict[str, Dict[str, Any]],
        background_tasks: BackgroundTasks
    ) -> Dict[str, Any]:
        """Handle web scraping task."""
        if not services["scraper"]["available"]:
            raise HTTPException(
                status_code=503,
                detail="Scraping service unavailable. Please check service configuration."
            )
        
        # Extract URL from params
        url = params[0] if params else None
        if not url:
            return {"error": "No URL provided for scraping"}
        
        # Clean and validate URL
        url = url.strip().strip("'\"")
        if not url.startswith(("http://", "https://")):
            url = f"https://{url}"
        
        # Perform scraping
        scrape_params = {
            "extract_text": True,
            "extract_links": False,
            "extract_images": True,
            "extract_tables": True,
            "output_format": "json",
        }
        
        result = await call_maybe_async(scraper_service.scrape_url, url, scrape_params)
        
        if result and result.get("status") == "success" and services["postgres"]["available"]:
            # Store in knowledge base
            content = result.get("content", {})
            page_text = content.get("text", "")
            
            if len(page_text) > 100:
                documents = [{
                    "content": page_text,
                    "url": url,
                    "title": content.get("title", f"Content from {url}"),
                    "format": "text/html",
                    "timestamp": datetime.now().isoformat(),
                    "source": "orchestrated_scrape",
                    "images": content.get("images", []),
                }]
                background_tasks.add_task(store_document_task, documents)
        
        return result
    
    async def _handle_search_task(self,params: List[str],services: Dict[str, Dict[str, Any]],original_query: str) -> Dict[str, Any]:
        """Handle search/retrieval task."""
        if not services["postgres"]["available"]:
            return {
                "error": "Vector database unavailable",
                "fallback": "Using AI-only response",
                "result": await self._ai_only_fallback(original_query, services)}
        search_query = params[0] if params else original_query
        # Perform vector search
        search_results = await call_maybe_async(
            postgres_service.search_documents,
            search_query,
            n_results=50)
        if not search_results and services["ai_service"]["available"]:
            return await self._ai_only_fallback(search_query, services)
        # Generate enhanced response
        context = [r.get("content", "") for r in search_results[:5]]
        
        if services["ai_service"]["available"]:
            enhanced = await ai_service.generate_enhanced_response(
                search_query,
                context,
                None
            )
            return {
                "answer": enhanced.get("text") if isinstance(enhanced, dict) else enhanced,
                "sources": search_results[:5],
                "confidence": enhanced.get("quality_score", 0.8) if isinstance(enhanced, dict) else 0.8
            }
        
        return {
            "sources": search_results[:5],
            "message": "AI service unavailable for enhanced response generation"
        }
    
    async def _handle_analyze_task(
        self,
        params: List[str],
        services: Dict[str, Dict[str, Any]],
        original_query: str
    ) -> Dict[str, Any]:
        """Handle analysis/summarization task."""
        if not services["ai_service"]["available"]:
            raise HTTPException(
                status_code=503,
                detail="AI service unavailable for analysis tasks"
            )
        
        subject = params[0] if params else original_query
        
        # Get context from knowledge base if available
        context = []
        if services["postgres"]["available"]:
            search_results = await call_maybe_async(
                postgres_service.search_documents,
                subject,
                n_results=10
            )
            context = [r.get("content", "") for r in search_results if r.get("content")]
        
        # Perform analysis
        analysis = await ai_service.generate_enhanced_response(
            original_query,
            context,
            "explanatory"
        )
        
        return {
            "analysis": analysis.get("text") if isinstance(analysis, dict) else analysis,
            "quality_score": analysis.get("quality_score", 0.0) if isinstance(analysis, dict) else 0.0,
            "sources_used": len(context)
        }
    
    async def _handle_bulk_task(
        self,
        params: List[str],
        services: Dict[str, Dict[str, Any]],
        background_tasks: BackgroundTasks
    ) -> Dict[str, Any]:
        """Handle bulk operations."""
        if not services["scraper"]["available"]:
            raise HTTPException(
                status_code=503,
                detail="Scraper service unavailable for bulk operations"
            )
        
        base_url = params[0] if params else None
        if not base_url:
            return {"error": "No base URL provided for bulk operation"}
        
        # Clean URL
        base_url = base_url.strip().strip("'\"")
        if not base_url.startswith(("http://", "https://")):
            base_url = f"https://{base_url}"
        
        # Discover URLs
        discovered_urls = await call_maybe_async(
            scraper_service.discover_url,
            base_url,
            max_depth=8,
            max_urls=500
        )
        
        if discovered_urls:
            background_tasks.add_task(
                enhanced_bulk_scrape_task,
                discovered_urls,
                auto_store=services["postgres"]["available"],
                max_depth=8,
                extract_images=True
            )
        
        return {
            "status": "started",
            "base_url": base_url,
            "discovered_urls": len(discovered_urls) if discovered_urls else 0,
            "urls_preview": discovered_urls[:10] if discovered_urls else [],
             "max_depth": 8,
        "estimated_completion_minutes": len(discovered_urls) * 0.5 if discovered_urls else 0
        }
    
    async def _handle_generic_task(
        self,
        query: str,
        services: Dict[str, Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Handle generic/unclassified task."""
        if services["ai_service"]["available"]:
            return await self._ai_only_fallback(query, services)
        
        return {
            "error": "Unable to process task - no services available",
            "query": query,
            "services_status": services
        }
    
    async def _ai_only_fallback(
        self,
        query: str,
        services: Dict[str, Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Fallback to AI-only response when services unavailable."""
        if not services["ai_service"]["available"]:
            raise HTTPException(
                status_code=503,
                detail="No services available to process this request"
            )
        
        response = await ai_service.generate_response(query, [])
        return {
            "answer": response,
            "source": "ai_only",
            "note": "Response generated without knowledge base context"
        }


# Global orchestration service instance
orchestration_service = OrchestrationService()


# ===================== Utility Functions =====================

def chunk_text(text: str, chunk_size: int = 1500, overlap: int = 200) -> List[str]:
    """Chunk text with sentence/paragraph boundaries"""
    if not text:
        return []
    if len(text) <= chunk_size:
        return [text]
    chunks = []
    start = 0
    text_len = len(text)
    while start < text_len:
        end = min(start + chunk_size, text_len)
        if end < text_len:
            paragraph_break = text.rfind("\n\n", start, end)
            if paragraph_break > start + (chunk_size // 2):
                end = paragraph_break
            else:
                sentence_break = text.rfind(". ", start, end)
                if sentence_break > start + (chunk_size // 2):
                    end = sentence_break + 1
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        next_start = end - overlap
        if next_start <= start:
            start = end
        else:
            start = next_start
    return chunks

def _extract_key_concepts(text: str) -> List[str]:
    """Extract key concepts from text for relevance matching."""
    if not text:
        return []
    
    # Technical patterns
    patterns = [
        r"\b[A-Z]{2,}\b",           # Acronyms (API, AWS, SQL)
        r"\b\w+(?:_\w+)+\b",        # Snake_case (node_red, python_3)
        r"\b\w+(?:-\w+)+\b",        # Kebab-case (docker-compose)
        r"\b\d+(?:\.\d+)?\w*\b"     # Versions (3.14, v2.0)
    ]
    
    concepts = []
    for p in patterns:
        concepts.extend(re.findall(p, text))
    
    # Extract meaningful words (5+ chars)
    words = re.findall(r"\b[a-zA-Z]{5,}\b", text)
    
    # Filter stopwords
    stopwords = {
        "about", "after", "again", "before", "being", "could", 
        "while", "would", "their", "there", "these", "those"
    }
    meaningful = [w.lower() for w in words if w.lower() not in stopwords]
    
    # Combine and deduplicate
    combined = list(dict.fromkeys(concepts + meaningful))
    return combined[:15]



async def call_maybe_async(fn, *args, **kwargs):
    """Call a function that might be sync or async."""
    if not callable(fn):
        raise RuntimeError("Provided object is not callable")
    try:
        result = fn(*args, **kwargs)
    except TypeError:
        result = fn(*args, **kwargs)
    if inspect.isawaitable(result):
        return await result
    return result


# ===================== API Endpoints =====================

@router.post("/widget/query")
async def widget_query(request: WidgetQueryRequest, background_tasks: BackgroundTasks, http_request: Request = None):
    """
    Query processing with image extraction
    
    User authentication is extracted from the Authorization header (Keycloak token).
    """
    try:
        query = request.query.strip()
        if not query:
            raise HTTPException(status_code=400, detail="Query cannot be empty")
        
        # Extract user_id from Authorization header if not provided in request body
        if not request.user_id and http_request:
            request.user_id = get_user_id_from_request(
                authorization=http_request.headers.get("Authorization"),
                x_user_id=http_request.headers.get("X-User-Id"),
                x_user_email=http_request.headers.get("X-User-Email"),
                default=None
            )
            if request.user_id:
                logger.info(f"📋 Extracted user_id from token: {request.user_id}")

        logger.info(f"Processing widget query: '{query}' (force_rag_only: {request.force_rag_only}, user: {request.user_id})")
        
        # Session management
        session_id = await _get_or_create_session_id(request)
        
        # Agent routing (if not force_rag_only)
        if request.force_rag_only:
            should_route_to_agent = False
        else:
            should_route_to_agent = await _should_route_to_agent(query, session_id)
        
        # ==================== STEP 3: Route to Agent Manager if Applicable ====================
        if should_route_to_agent:
            logger.info(f"🎯 Routing to Agent Manager")
            # Extract auth token from Authorization header
            auth_token = None
            user_type = None
            if http_request:
                auth_header = http_request.headers.get("Authorization", "")
                if auth_header.startswith("Bearer "):
                    auth_token = auth_header.replace("Bearer ", "")
                # Extract user type (ENG = Engineer, CUS = Customer)
                user_type = http_request.headers.get("usertype") or http_request.headers.get("Usertype")
                if user_type:
                    logger.info(f"👤 User type from header: {user_type}")
            return await _handle_agent_routing(
                query=query,
                session_id=session_id,
                request=request,
                background_tasks=background_tasks,
                auth_token=auth_token,
                user_type=user_type
            )
        
        #  Standard RAG flow with fast search
        search_results = await _perform_document_search(query, request)
        
        if not search_results:
            return await _handle_no_results(query, {}, request)

        #  Early filtering
        filtered_results = _filter_search_results(search_results, request)
        base_context = _extract_base_context(filtered_results)
        
        if not base_context:
            base_context = [r.get("content", "")[:1000] for r in filtered_results[:3]]

        # Generate enhanced response
        answer, expanded_context, confidence = await _generate_enhanced_response(
            query, base_context
        )

        # Generate steps
        steps_data = await _generate_steps(query, expanded_context, base_context, answer)

        #  Fast parallel image extraction
        selected_images = _extract_and_score_images(
            query=query,
            answer=answer,
            filtered_results=filtered_results
        )

        # Build sources
        sources = _build_sources(filtered_results, request)

        # Generate summary
        summary = await _generate_summary(answer, expanded_context)

        # Combine steps with images
        steps_with_images = _combine_steps_with_images(steps_data, selected_images)

        # Calculate confidence
        final_confidence = _calculate_final_confidence(
            confidence, filtered_results, answer, request
        )

        # Store interaction
        if request.store_interaction:
            await _store_interaction(
                query=query,
                answer=answer,
                confidence=final_confidence,
                sources=sources,
                background_tasks=background_tasks
            )

        # Format response
        from app.services.webui_formatter import format_for_webui
        
        formatted_answer = format_for_webui(
            answer=answer or "I was unable to generate a comprehensive answer.",
            steps=steps_with_images,
            images=selected_images,
            query=query,
            confidence=final_confidence,
            summary=summary,
            metadata={
                "results_found": len(search_results),
                "results_used": len(filtered_results),
                "search_depth": request.search_depth
            }
        )

        return {
            "query": query,
            "answer": formatted_answer,
            "steps": steps_with_images,
            "images": selected_images,
            "sources": sources,
            "has_sources": len(sources) > 0,
            "confidence": round(final_confidence, 3),
            "search_depth": request.search_depth,
            "results_found": len(search_results),
            "results_used": len(filtered_results),
            "timestamp": datetime.now().isoformat(),
            "summary": summary,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Widget query error: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

# ==================== HELPER FUNCTIONS ====================

async def _get_or_create_session_id(request: WidgetQueryRequest) -> str:
    """Get or create session ID for conversation continuity."""
    import hashlib
    from datetime import datetime
    
    if request.session_id:
        logger.info(f"📋 Using provided session ID: {request.session_id}")
        return request.session_id
    
    time_bucket = str(datetime.now().hour) + str(datetime.now().minute // 10)
    
    if request.user_id:
        session_id = hashlib.md5(f"widget_{request.user_id}_{time_bucket}".encode()).hexdigest()[:16]
        logger.info(f"📋 Generated user session ID: {session_id} (10-min bucket)")
    else:
        session_id = hashlib.md5(f"widget_anon_{time_bucket}".encode()).hexdigest()[:16]
        logger.info(f"📋 Generated anonymous session ID: {session_id} (10-min bucket)")
    
    return session_id


async def _should_route_to_agent(query: str, session_id: str) -> bool:
    """
    Determine if query should be routed to agent manager.
    
    Strategy: Route ALL queries to the agent manager (orchestrator).
    The orchestrator intelligently handles:
    - Greetings ("hi", "hello")
    - Capability questions ("what can you do?")
    - Resource operations ("list clusters")
    - Documentation queries (routes to RAG internally)
    
    This ensures consistent, intelligent handling of all user inputs.
    """
    from app.agents.state.conversation_state import conversation_state_manager, ConversationStatus
    
    query_lower = query.lower().strip()
    
    # Skip empty queries
    if not query_lower:
        return False
    
    # Skip OpenWebUI metadata generation requests (title, tags, etc.)
    # These are internal requests that should go to RAG/AI directly
    metadata_patterns = [
        "### task:",
        "generate a concise",
        "suggest 3-5 relevant follow-up",
        "generate 1-3 broad tags"
    ]
    if any(pattern in query_lower for pattern in metadata_patterns):
        logger.info(f"📋 Metadata request detected - skipping agent routing")
        return False
    
    # Route EVERYTHING else to the agent manager
    # The orchestrator will intelligently decide:
    # - Greetings → greeting response
    # - Capability questions → capability response
    # - Resource operations → intent agent → execution
    # - Documentation questions → RAG agent
    logger.info(f"🎯 Routing to agent manager for intelligent handling: '{query[:50]}...'")
    return True


async def _llm_intent_classification(query: str) -> bool:
    """Use LLM to classify query intent and detect typos."""
    logger.info(f"🤖 Using LLM to analyze query intent: '{query}'")
    try:
        intent_prompt = f"""You are an intent classifier for a cloud infrastructure management system.
User query: "{query}"
Analyze if this query is about:
1. Managing cloud resources (clusters, endpoints, datacenters, firewalls, databases, etc.)
2. Resource operations (create, list, show, get, delete, update, etc.)

Even if the user has typos (e.g., "lis" instead of "list"), understand the intent.

Respond ONLY with a JSON object:
{{
  "is_resource_operation": true/false,
  "corrected_query": "the query with typos fixed",
  "action": "create/list/show/delete/update/etc or null",
  "resource": "cluster/endpoint/firewall/etc or null",
  "confidence": 0.0-1.0
}}"""
        
        llm_response = await ai_service._call_chat_with_retries(
            prompt=intent_prompt,
            max_tokens=200,
            temperature=0.1,
            timeout=15
        )
        
        import json
        import re
        
        json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', llm_response)
        if json_match:
            intent_data = json.loads(json_match.group(0))
            
            is_resource_op = intent_data.get("is_resource_operation", False)
            confidence = intent_data.get("confidence", 0.0)
            
            logger.info(f"🤖 LLM Intent: is_resource_op={is_resource_op}, confidence={confidence:.0%}")
            
            if is_resource_op and confidence >= 0.7:
                logger.info(f"✅ LLM confirmed resource operation")
                return True
        
        return False
        
    except Exception as e:
        logger.error(f"❌ LLM intent classification failed: {str(e)}")
        return False


def _check_conversation_context(existing_state, query_lower: str) -> bool:
    """Check if query continues an existing conversation."""
    if len(existing_state.conversation_history) == 0:
        return False
    
    last_messages = [msg for msg in existing_state.conversation_history 
                    if msg.get("role") == "assistant"]
    
    if not last_messages:
        return False

    last_response = last_messages[-1].get("content", "").lower()
    conversation_indicators = ["cluster", "data center", "endpoint", "which one"]
    if any(word in last_response for word in conversation_indicators):
        logger.info(f"🎯 Query continues conversation context → routing to agents")
        return True
    return False

async def _handle_agent_routing(query: str, session_id: str, request: WidgetQueryRequest, 
                                background_tasks: BackgroundTasks, auth_token: str = None,
                                user_type: str = None) -> dict:
    """Handle routing to agent manager and process response."""
    from app.services.webui_formatter import format_agent_response_for_webui
    
    agent_manager = get_agent_manager(
        vector_service=postgres_service,
        ai_service=ai_service)
    
    user_roles = getattr(request, 'user_roles', None) or ["admin", "developer", "viewer"]
    agent_result = await agent_manager.process_request(
        user_input=query,
        session_id=session_id,
        user_id=request.user_id or "widget_user",
        user_roles=user_roles,
        auth_token=auth_token,
        user_type=user_type)
    
    logger.info(f"✅ Agent processing complete: routing={agent_result.get('routing')}, "
               f"success={agent_result.get('success')}")

    response_text = agent_result.get("response", "")
    metadata = agent_result.get("metadata", {})
    follow_ups = agent_result.get("follow_ups", [])  # Get follow-ups from orchestrator
    missing_params = metadata.get("missing_params", [])
    
    if missing_params or "?" in response_text or "which" in response_text.lower():
        logger.info(f"🔄 Agent asking for clarification, missing: {missing_params}")
        return {
            "query": query,
            "answer": response_text,
            "sources": [],
            "intent_detected": True,
            "routed_to": "agent_manager",
            "conversation_active": True,
            "session_id": session_id,
            "missing_params": missing_params,
            "metadata": metadata,
            "timestamp": datetime.now().isoformat(),
            "results_found": 100,
            "confidence": 0.95,
            "has_sources": True,
            "images": [],
            "steps": [],
            "followUps": follow_ups  # Include follow-ups for OpenWebUI
        }

    execution_result = agent_result.get("execution_result")
    if execution_result and execution_result.get("success"):
        logger.info(f"🎯 Agent executed operation successfully")
        result_data = execution_result.get("data", {})

        if isinstance(result_data, dict) and "endpoints" in result_data:
            resp = _format_endpoint_response(result_data, query, session_id, metadata)
            resp["followUps"] = follow_ups
            return resp

        if isinstance(result_data, dict) and "data" in result_data:
            resp = _format_cluster_response(result_data, query, session_id, metadata)
            resp["followUps"] = follow_ups
            return resp

    formatted_answer = format_agent_response_for_webui(
        response_text=response_text,
        execution_result=execution_result,
        session_id=session_id,
        metadata=metadata
    )
    return {
        "query": query,
        "answer": formatted_answer,
        "sources": [],
        "intent_detected": True,
        "routed_to": "agent_manager",
        "session_id": session_id,
        "metadata": metadata,
        "timestamp": datetime.now().isoformat(),
        "results_found": 100,
        "confidence": 0.99,
        "has_sources": True,
        "images": [],
        "steps": [],
        "followUps": follow_ups  # Include follow-ups for OpenWebUI
    }

def _format_endpoint_response(result_data: dict, query: str, session_id: str, metadata: dict) -> dict:
    """Format endpoint listing response."""
    endpoints = result_data.get("endpoints", [])
    total = result_data.get("total", len(endpoints))
    
    answer = f"📍 **Available Endpoints/Datacenters** ({total} found)\n\n"
    answer += "| # | Name | ID | Type |\n"
    answer += "|---|------|----|----- |\n"
    
    for i, ep in enumerate(endpoints, 1):
        name = ep.get("name", "Unknown")
        ep_id = ep.get("id", "N/A")
        ep_type = ep.get("type", "")
        answer += f"| {i} | {name} | {ep_id} | {ep_type} |\n"
    
    answer += "\n💡 You can use these endpoints when listing or creating clusters."
    return {
        "query": query,
        "answer": answer,
        "sources": [],
        "intent_detected": True,
        "routed_to": "agent_manager",
        "auto_executed": True,
        "session_id": session_id,
        "execution_result": {
            "total_endpoints": total,
            "endpoints": endpoints
        },
        "metadata": metadata,
        "timestamp": datetime.now().isoformat(),
        "results_found": total,
        "results_used": total,
        "confidence": 0.99,
        "has_sources": True,
        "images": [],
        "steps": []
    }

def _format_cluster_response(result_data: dict, query: str, session_id: str, metadata: dict) -> dict:
    """Format cluster listing response."""
    clusters = result_data["data"]

    by_endpoint = {}
    for cluster in clusters:
        endpoint = cluster.get("displayNameEndpoint", "Unknown")
        if endpoint not in by_endpoint:
            by_endpoint[endpoint] = []
        by_endpoint[endpoint].append(cluster)

    if len(by_endpoint) == 1:
        endpoint_name = list(by_endpoint.keys())[0]
        answer = f"✅ Found **{len(clusters)} Kubernetes clusters** in **{endpoint_name}**:\n\n"
    else:
        answer = f"✅ Found **{len(clusters)} Kubernetes clusters** across **{len(by_endpoint)} data centers**:\n\n"
    
    for endpoint, endpoint_clusters in sorted(by_endpoint.items()):
        answer += f"📍 **{endpoint}** ({len(endpoint_clusters)} clusters)\n"
        for cluster in endpoint_clusters:
            status_emoji = "✅" if cluster.get("status") == "Healthy" else "⚠️"
            answer += f"  {status_emoji} {cluster.get('clusterName', 'N/A')} - "
            answer += f"{cluster.get('nodescount', 0)} nodes, "
            answer += f"K8s {cluster.get('kubernetesVersion', 'N/A')}\n"
        answer += "\n"
    
    return {
        "query": query,
        "answer": answer,
        "sources": [],
        "intent_detected": True,
        "routed_to": "agent_manager",
        "auto_executed": True,
        "session_id": session_id,
        "execution_result": {
            "total_clusters": len(clusters),
            "endpoints": len(by_endpoint),
            "clusters": clusters
        },
        "metadata": metadata,
        "timestamp": datetime.now().isoformat(),
        "results_found": len(clusters),
        "results_used": len(clusters),
        "confidence": 0.99,
        "has_sources": True,
        "images": [],
        "steps": []
    }


async def _handle_auto_execution(query: str, intent_analysis: dict, request: WidgetQueryRequest,background_tasks: BackgroundTasks) -> dict:
    """Handle auto-execution of detected tasks."""
    primary_task = intent_analysis.get("primary_task")
    if not primary_task or primary_task["confidence"] <= 0.7:
        return None
    logger.info(f"Auto-executing task: {primary_task['type']}")
    execution_result = await orchestration_service.execute_orchestrated_task(
        primary_task,
        background_tasks)
    if request.store_interaction:
        interaction_doc = {
            "content": json.dumps({
                "query": query,
                "intent": intent_analysis,
                "execution": execution_result
            }),
            "url": f"interaction://{datetime.now().timestamp()}",
            "title": f"User Interaction: {query[:50]}",
            "format": "application/json",
            "timestamp": datetime.now().isoformat(),
            "source": "widget_interaction",
        }
        background_tasks.add_task(store_document_task, [interaction_doc])
    return {
        "query": query,
        "intent_detected": intent_analysis,
        "auto_executed": True,
        "execution_result": execution_result,
        "timestamp": datetime.now().isoformat()
    }

async def _perform_document_search(query: str, request: WidgetQueryRequest) -> list:
    """
     Fast document search with reduced candidates
    """
    search_params = {
        "quick": {"max_results": min(request.max_results, 20)},      # Reduced from 30
        "balanced": {"max_results": min(request.max_results, 40)},   # Reduced from 50
        "deep": {"max_results": min(request.max_results, 60)},       # Reduced from 100
    }
    
    search_config = search_params.get(request.search_depth, search_params["balanced"])
    
    try:
        search_results = await postgres_service.search_documents(
            query,
            n_results=search_config["max_results"]
        )
        return search_results or []
    except Exception as e:
        logger.exception(f"Error while searching documents: {e}")
        return []
    
async def _handle_no_results(query: str, intent_analysis: dict, request: WidgetQueryRequest) -> dict:
    """Handle case when no search results are found."""
    try:
        answer = await call_maybe_async(ai_service.generate_response, query, [])
    except Exception as e:
        logger.warning(f"LLM fallback failed: {e}")
        answer = None
    if answer:
        summary = None
        try:
            summary = await call_maybe_async(
                ai_service.generate_summary,
                answer,
                max_sentences=3,
                max_chars=600
            )
        except Exception:
            summary = (answer[:600] + "...") if len(answer) > 600 else answer
        return {
            "query": query,
            "answer": answer,
            "intent_detected": intent_analysis,
            "auto_executed": False,
            "steps": [],
            "images": [],
            "sources": [],
            "has_sources": False,
            "confidence": 0.45,
            "search_depth": request.search_depth,
            "timestamp": datetime.now().isoformat(),
            "summary": summary or "No summary available."
        }
    return {
        "query": query,
        "answer": "I don't have any relevant information in my knowledge base to answer your question. Please try rephrasing your query or add more context.",
        "intent_detected": intent_analysis,
        "steps": [],
        "images": [],
        "sources": [],
        "has_sources": False,
        "confidence": 0.0,
        "search_depth": request.search_depth,
        "timestamp": datetime.now().isoformat()
    }

def _filter_search_results(search_results: list, request: WidgetQueryRequest) -> list:
    """
     Early filtering with adaptive thresholds
    """
    if not search_results:
        return []
    
    # Calculate adaptive threshold
    avg_score = sum(r.get("relevance_score", 0) for r in search_results) / max(1, len(search_results))
    min_threshold = max(0.15, avg_score * 0.5)  # More permissive threshold
    
    # Filter by threshold
    filtered = [r for r in search_results if r.get("relevance_score", 0) >= min_threshold]
    
    # Ensure minimum results
    if len(filtered) < 3:
        filtered = search_results[:max(5, request.max_results // 2)]
    
    logger.info(
        f"📊 Filtered {len(search_results)} → {len(filtered)} results "
        f"(threshold: {min_threshold:.2f})"
    )
    
    return filtered
def _extract_base_context(filtered_results: list) -> list:
    """Extract base context from filtered results."""
    base_context = []
    for result in filtered_results:
        content = result.get("content", "") if isinstance(result, dict) else ""
        if content and len(content.strip()) > 50:
            score = result.get("relevance_score", 0.5)
            if score > 0.7:
                base_context.append(content)
            elif score > 0.5:
                base_context.append(content[:1500])
            else:
                base_context.append(content[:800])
    
    return base_context

async def _generate_enhanced_response(query: str, base_context: list) -> tuple:
    """Generate enhanced response with expanded context."""
    try:
        enhanced_result = await call_maybe_async(
            ai_service.generate_enhanced_response,
            query,
            base_context,
            None
        )
        if isinstance(enhanced_result, dict):
            answer = enhanced_result.get("text", "")
            expanded_context = enhanced_result.get("expanded_context", "")
            confidence = enhanced_result.get("quality_score", 0.0)
        else:
            answer = enhanced_result or ""
            expanded_context = ""
            confidence = 0.0   
    except Exception as e:
        logger.warning(f"Enhanced response generation failed: {e}")
        try:
            answer = await call_maybe_async(ai_service.generate_response, query, base_context[:3])
        except Exception as e2:
            logger.error(f"Fallback generate_response also failed: {e2}")
            answer = ""
        expanded_context = "\n\n".join(base_context[:2]) if base_context else ""
        confidence = 0.6
    return answer, expanded_context, confidence

async def _generate_steps(query: str, expanded_context: str, base_context: list, answer: str) -> list:
    """Generate stepwise response."""
    working_context = [expanded_context] if expanded_context else base_context[:3]
    try:
        steps_data = await call_maybe_async(
            ai_service.generate_stepwise_response,
            query,
            working_context
        )
    except Exception as e:
        logger.warning(f"Stepwise generation failed: {e}")
        steps_data = []

    if not steps_data:
        if answer:
            sentences = [s.strip() for s in answer.split(".") if s.strip()]
            steps_data = [{"text": (s + "."), "type": "info"} for s in sentences[:5]]
        else:
            steps_data = [{"text": "Unable to generate structured response.", "type": "info"}]
    return steps_data

def _extract_and_score_images(
    query: str,
    answer: str,
    filtered_results: list
) -> list:
    """
     Parallel image extraction with  relevance scoring
    """
    import concurrent.futures
    from collections import defaultdict
    
    if not filtered_results:
        return []
    
    # Extract query concepts once
    query_concepts = set(_extract_key_concepts(query.lower()))
    answer_concepts = set(_extract_key_concepts(answer.lower())) if answer else set()
    all_concepts = query_concepts | answer_concepts
    
    # Track query keywords for relevance
    query_keywords = set(re.findall(r'\b\w{4,}\b', query.lower()))
    
    logger.info(f"🔍 Extracting images from {len(filtered_results)} results")
    logger.debug(f"Query concepts: {list(all_concepts)[:10]}")

    def process_result(result_data):
        """Process single result in parallel."""
        result_idx, result = result_data
        images_from_result = []
        
        meta = result.get("metadata", {}) or {}
        page_url = meta.get("url", "")
        page_title = meta.get("title", "")
        relevance_score = result.get("relevance_score", 0.0)
        
        # Extract images
        images_raw = meta.get("images", [])
        if not images_raw:
            images_raw = meta.get("images_json", [])
        
        if not isinstance(images_raw, list):
            if isinstance(images_raw, str):
                try:
                    images_raw = json.loads(images_raw)
                except:
                    return images_from_result
            else:
                return images_from_result
        
        if not images_raw:
            return images_from_result
        
        # Process each image
        for img in images_raw:
            if not isinstance(img, dict):
                if isinstance(img, str) and img.startswith("http"):
                    img = {"url": img, "alt": "", "caption": "", "type": "content"}
                else:
                    continue
            
            img_url = img.get("url", "")
            if not img_url or not isinstance(img_url, str):
                continue
            
            #  Early URL validation
            if not img_url.startswith(("http://", "https://")):
                continue
            
            # Filter low-quality images
            skip_patterns = [
                "1x1", "pixel", "tracker", "blank", "spacer",
                "loading", "spinner", "placeholder", "icon-", 
                "logo-small", "avatar-", "thumb-"
            ]
            if any(skip in img_url.lower() for skip in skip_patterns):
                continue
            
            # Extract metadata
            img_alt = img.get("alt", "")
            img_caption = img.get("caption", "")
            img_type = img.get("type", "content")
            img_text = img.get("text", "")
            
            #   relevance calculation
            img_search_text = " ".join([
                img_alt, img_caption, img_text, page_title
            ]).lower()
            
            # Calculate relevance components
            img_concepts = set(_extract_key_concepts(img_search_text))
            concept_overlap = len(all_concepts & img_concepts) if all_concepts and img_concepts else 0
            
            # Keyword matching
            img_keywords = set(re.findall(r'\b\w{4,}\b', img_search_text))
            keyword_overlap = len(query_keywords & img_keywords) if query_keywords else 0
            
            # Type-based scoring
            type_bonus = 0.0
            if img_type in ["diagram", "screenshot", "illustration", "chart"]:
                type_bonus = 0.3
            elif img_type in ["photo", "image"]:
                type_bonus = 0.1
            
            # Quality score from existing metadata
            quality_score = float(img.get("quality_score", 0.5))
            
            #  Combined relevance score
            image_score = (
                (concept_overlap / max(len(all_concepts), 1)) * 0.30 +  # Concept match
                (keyword_overlap / max(len(query_keywords), 1)) * 0.25 + # Keyword match
                relevance_score * 0.20 +                                  # Parent doc relevance
                type_bonus * 0.15 +                                       # Image type
                quality_score * 0.10                                      # Image quality
            )
            
            # Quality threshold
            if image_score > 0.10:  # Keep threshold low for better recall
                images_from_result.append({
                    "url": img_url,
                    "alt": img_alt or "Image",
                    "type": img_type,
                    "caption": img_caption,
                    "source_url": page_url,
                    "source_title": page_title,
                    "relevance_score": round(image_score, 3),
                    "text": img_text[:500],
                    "concept_overlap": concept_overlap,
                    "keyword_overlap": keyword_overlap,
                })
        
        return images_from_result
    
    #  Process results in parallel
    candidate_images = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        # Submit all processing tasks
        future_to_result = {
            executor.submit(process_result, (idx, result)): idx
            for idx, result in enumerate(filtered_results)
        }
        
        # Collect results as they complete
        for future in concurrent.futures.as_completed(future_to_result):
            try:
                images = future.result(timeout=5)
                candidate_images.extend(images)
            except Exception as e:
                logger.debug(f"Image processing error: {e}")
    
    #   deduplication (keep highest scored)
    seen_urls = {}
    
    for img in candidate_images:
        url = img.get("url")
        score = img.get("relevance_score", 0)
        
        if url not in seen_urls or seen_urls[url]["relevance_score"] < score:
            seen_urls[url] = img
    
    # Sort by relevance score
    unique_images = sorted(
        seen_urls.values(),
        key=lambda x: x.get("relevance_score", 0),
        reverse=True
    )
    
    logger.info(
        f"✅ Extracted {len(unique_images)} unique images "
        f"(from {len(candidate_images)} candidates)"
    )
    
    # Log top images
    if unique_images:
        logger.info("📊 Top 3 images:")
        for i, img in enumerate(unique_images[:3], 1):
            logger.info(
                f"  {i}. {img.get('url', 'N/A')[:60]} "
                f"(score: {img.get('relevance_score', 0):.3f}, "
                f"concepts: {img.get('concept_overlap', 0)}, "
                f"keywords: {img.get('keyword_overlap', 0)})"
            )
    
    return unique_images[:12]



def _build_sources(filtered_results: list, request: WidgetQueryRequest) -> list:
    """Build sources list from filtered results."""
    sources = []
    if request.include_sources:
        for result in filtered_results:
            meta = result.get("metadata", {}) or {}
            content_preview_raw = result.get("content", "") or ""
            content_preview = (
                content_preview_raw[:300] + "..." if len(content_preview_raw) > 300 else content_preview_raw
            )
            sources.append({
                "url": meta.get("url", ""),
                "title": meta.get("title", "Untitled"),
                "relevance_score": round(result.get("relevance_score", 0), 3),
                "content_preview": content_preview,
                "domain": meta.get("domain", ""),
                "last_updated": meta.get("timestamp", ""),
            })
    
    return sources


async def _generate_summary(answer: str, expanded_context: str) -> str:
    """Generate summary from answer or context."""
    summary_input = answer if answer else expanded_context
    
    if summary_input:
        try:
            summary = await call_maybe_async(
                ai_service.generate_summary,
                summary_input,
                max_sentences=4,
                max_chars=600
            )
        except Exception:
            summary = summary_input[:600] + "..." if len(summary_input) > 600 else summary_input
    else:
        summary = "No summary available."
    
    return summary


def _combine_steps_with_images(steps_data: list, selected_images: list) -> list:
    """Combine steps with relevant images."""
    steps_with_images = []
    
    for i, step in enumerate(steps_data):
        step_obj = {
            "index": i + 1,
            "text": step.get("text", ""),
            "type": step.get("type", "action")
        }
        assigned_img = None

        if isinstance(step, dict):
            # Check for existing image in step
            si = step.get("image") if isinstance(step.get("image"), (dict, str)) else None
            
            if isinstance(si, dict) and si.get("url"):
                assigned_img = {
                    "url": si.get("url"),
                    "alt": si.get("alt", "") or step.get("alt", "") or "",
                    "caption": si.get("caption", "") or step.get("caption", "") or "",
                    "relevance_score": si.get("relevance_score", None),
                }
            elif isinstance(si, str) and si.startswith("http"):
                assigned_img = {
                    "url": si,
                    "alt": step.get("alt", "") or "",
                    "caption": step.get("caption", "") or "",
                    "relevance_score": None
                }

            if not assigned_img:
                image_prompt = None
                if isinstance(step.get("image_prompt"), str) and step.get("image_prompt").strip():
                    image_prompt = step.get("image_prompt").strip()
                elif isinstance(step.get("image"), dict) and isinstance(step["image"].get("image_prompt"), str):
                    image_prompt = step["image"].get("image_prompt").strip()
                elif isinstance(step.get("image"), str) and not step.get("image").startswith("http") and len(step.get("image").strip()) > 0:
                    image_prompt = step.get("image").strip()

                if image_prompt:
                    assigned_img = {"image_prompt": image_prompt}

        if not assigned_img and selected_images and i < len(selected_images):
            step_img = selected_images[i]
            if step_img.get("url"):
                assigned_img = {
                    "url": step_img.get("url"),
                    "alt": step_img.get("alt", "") or "",
                    "caption": step_img.get("caption", "") or "",
                    "relevance_score": step_img.get("relevance_score"),
                }

        if assigned_img:
            step_obj["image"] = assigned_img

        steps_with_images.append(step_obj)
    
    return steps_with_images


def _calculate_final_confidence(confidence: float, filtered_results: list, 
                                answer: str, request: WidgetQueryRequest) -> float:
    """Calculate final confidence score."""
    return min(
        1.0,
        (
            (confidence or 0.0) * 0.4
            + (len(filtered_results) / max(request.max_results, 10)) * 0.3
            + (1.0 if answer and len(answer) > 100 else 0.5) * 0.3
        ),
    )


async def _store_interaction(query: str, answer: str, confidence: float, 
                             sources: list, background_tasks: BackgroundTasks):
    """Store interaction for future reference."""
    interaction_doc = {
        "content": json.dumps({
            "query": query,
            "answer": answer,
            "confidence": confidence,
            "sources_count": len(sources)
        }),
        "url": f"interaction://{datetime.now().timestamp()}",
        "title": f"Query: {query[:50]}",
        "format": "application/json",
        "timestamp": datetime.now().isoformat(),
        "source": "widget_query",
    }
    background_tasks.add_task(store_document_task, [interaction_doc])

def _enhanced_similarity(query: str, text: str) -> float:
    """Calculate enhanced similarity between query and text"""
    if not query or not text:
        return 0.0
    
    query_norm = " ".join(query.lower().split())
    text_norm = " ".join(text.lower().split())
    
    # Sequence matching
    import difflib
    seq_ratio = difflib.SequenceMatcher(None, query_norm, text_norm).ratio()
    # Word overlap
    query_words = set(query_norm.split())
    text_words = set(text_norm.split())
    overlap = len(query_words & text_words) / len(query_words | text_words) if query_words else 0.0
    # Substring bonus
    substring_bonus = 0.2 if query_norm in text_norm else 0.0
    
    score = 0.4 * seq_ratio + 0.4 * overlap + 0.2 * substring_bonus
    return min(1.0, max(0.0, score))

@router.post("/widget/execute-task")
async def widget_execute_task(request: TaskExecutionRequest, background_tasks: BackgroundTasks):
    """
    Execute arbitrary task with automatic service orchestration.
    Analyzes task requirements and connects appropriate services.
    """
    try:
        logger.info(f"Executing task: {request.task_description[:100]}")
        
        # Detect task intent
        intent_analysis = await orchestration_service.detect_task_intent(request.task_description)
        
        if not intent_analysis.get("primary_task"):
            if request.store_result:
                unclear_doc = {
                    "content": json.dumps({
                        "task": request.task_description,
                        "status": "unclear",
                        "analysis": intent_analysis,
                        "context": request.context
                    }),
                    "url": f"task://{datetime.now().timestamp()}",
                    "title": f"Unclear Task: {request.task_description[:50]}",
                    "format": "application/json",
                    "timestamp": datetime.now().isoformat(),
                    "source": "unclear_task",
                }
                background_tasks.add_task(store_document_task, [unclear_doc])
            
            return {
                "status": "needs_clarification",
                "task": request.task_description,
                "message": "I need more information to complete this task. Could you please provide more details?",
                "suggestions": [
                    "What specific action would you like me to perform?",
                    "Are you looking to search, scrape, analyze, or something else?",
                    "Please provide any relevant URLs, files, or context."
                ],
                "intent_analysis": intent_analysis,
                "timestamp": datetime.now().isoformat()
            }
        
        # Check services if auto-connect enabled
        if request.auto_connect_services:
            services = await orchestration_service.check_service_availability()
            
            # Verify required services are available
            primary_task = intent_analysis["primary_task"]
            task_type = primary_task["type"]
            
            required_services = {
                "scrape": ["scraper"],
                "search": ["postgres", "ai_service"],
                "analyze": ["ai_service"],
                "bulk_operation": ["scraper", "postgres"]
            }
            
            missing_services = []
            for service_name in required_services.get(task_type, []):
                if not services.get(service_name, {}).get("available"):
                    missing_services.append(service_name)
            
            if missing_services:
                error_msg = f"Required services unavailable: {', '.join(missing_services)}"
                logger.warning(error_msg)
                # Store failed attempt
                if request.store_result:
                    failed_doc = {
                        "content": json.dumps({
                            "task": request.task_description,
                            "status": "failed",
                            "error": error_msg,
                            "missing_services": missing_services,
                            "services_status": services
                        }),
                        "url": f"task://{datetime.now().timestamp()}",
                        "title": f"Failed Task: {request.task_description[:50]}",
                        "format": "application/json",
                        "timestamp": datetime.now().isoformat(),
                        "source": "failed_task",
                    }
                    background_tasks.add_task(store_document_task, [failed_doc])
                return {
                    "status": "failed",
                    "task": request.task_description,
                    "error": error_msg,
                    "missing_services": missing_services,
                    "services_status": services,
                    "timestamp": datetime.now().isoformat()
                }
        
        # Execute task
        execution_result = await orchestration_service.execute_orchestrated_task(
            intent_analysis["primary_task"],
            background_tasks)
        # Store result if enabled
        if request.store_result and execution_result.get("status") == "success":
            result_doc = {
                "content": json.dumps({
                    "task": request.task_description,
                    "execution": execution_result,
                    "context": request.context
                }),
                "url": f"task://{datetime.now().timestamp()}",
                "title": f"Completed Task: {request.task_description[:50]}",
                "format": "application/json",
                "timestamp": datetime.now().isoformat(),
                "source": "completed_task",
            }
            background_tasks.add_task(store_document_task, [result_doc])
        
        return {
            "status": execution_result.get("status", "unknown"),
            "task": request.task_description,
            "intent_analysis": intent_analysis,
            "execution_result": execution_result,
            "timestamp": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Task execution error: {e}")
        raise HTTPException(status_code=500, detail=f"Task execution error: {str(e)}")

@router.post("/widget/scrape")
async def widget_scrape(request: WidgetScrapeRequest, background_tasks: BackgroundTasks):
    """
    Enhanced scraping with proper image extraction and storage.
    """
    try:
        logger.info(f"🌐 Scraping URL with image extraction: {request.url}")
        
        scrape_params = {
            "extract_text": True,
            "extract_links": False,
            "extract_images": True, 
            "extract_tables": True,
            "scroll_page": request.wait_for_js,
            "wait_for_element": "body" if request.wait_for_js else None,
            "output_format": "json",
        }
        # Scrape the page
        result = await call_maybe_async(
            scraper_service.scrape_url, 
            str(request.url), 
            scrape_params
        )
        
        if not result or result.get("status") != "success":
            raise HTTPException(
                status_code=400,
                detail=f"Scraping failed: {result.get('error', 'Unknown error')}"
            )
        content = result.get("content", {}) or {}
        page_text = content.get("text", "") or ""
        if len(page_text.strip()) < 50:
            raise HTTPException(
                status_code=400, 
                detail="Scraped content is too short or empty"
            )
        #  Extract images from scraped content
        scraped_images = content.get("images", []) or []
        logger.info(
            f"✅ Scraped {len(page_text)} chars, "
            f"{len(scraped_images)} images from {request.url}"
        )

        #  Store in knowledge base with images
        if request.store_in_knowledge:
            documents_to_store = []
            
            if len(page_text) > 2000:
                # Split into chunks
                chunks = chunk_text(page_text, chunk_size=1500, overlap=200)
                for i, chunk in enumerate(chunks):
                    chunk_images = scraped_images if i == 0 else []
                    documents_to_store.append({
                        "content": chunk,
                        "url": f"{str(request.url)}#chunk-{i}",
                        "title": content.get("title", "") or f"Content from {request.url}",
                        "format": "text/html",
                        "timestamp": datetime.now().isoformat(),
                        "source": "widget_scrape",
                        "images": chunk_images,  })
            else:
                # Single document with all images
                documents_to_store.append({
                    "content": page_text,
                    "url": str(request.url),
                    "title": content.get("title", "") or f"Content from {request.url}",
                    "format": "text/html",
                    "timestamp": datetime.now().isoformat(),
                    "source": "widget_scrape",
                    "images": scraped_images,  
                })

            # Store in background
            background_tasks.add_task(store_document_task, documents_to_store)
            
            logger.info(
                f"✅ Queued {len(documents_to_store)} documents for storage "
                f"with {len(scraped_images)} images"
            )
        # Generate summary
        try:
            summary = await call_maybe_async(
                ai_service.generate_summary,
                page_text,
                max_sentences=4,
                max_chars=800
            )
        except Exception:
            summary = page_text[:800] + "..." if len(page_text) > 800 else page_text
        #  Include images in response
        return {
            "status": "success",
            "url": str(request.url),
            "title": content.get("title", "Untitled"),
            "content_length": len(page_text),
            "word_count": len(page_text.split()),
            "images_count": len(scraped_images),  
            "images": scraped_images[:10],  
            "tables_count": len(content.get("tables", [])),
            "method_used": result.get("method"),
            "stored_in_knowledge": request.store_in_knowledge,
            "chunks_created": len(chunk_text(page_text)) if len(page_text) > 2000 else 1,
            "timestamp": result.get("timestamp"),
            "summary": summary,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Widget scrape error: {e}")
        raise HTTPException(status_code=500, detail=f"Scraping error: {str(e)}")
    
@router.post("/widget/bulk-scrape")
async def widget_bulk_scrape(request: BulkScrapeRequest, background_tasks: BackgroundTasks):
    """
    Enhanced bulk scraping with deeper crawling capabilities.
    New features:
    - Increased default depth from 5→8 (max 20)
    - Increased default max_urls from 200→500 (max 3000)
    - Option to follow external links
    - Deep image extraction from all pages
    """
    try:
        logger.info(
            f"🌐 Starting DEEP bulk scrape from: {request.base_url} "
            f"(depth={request.max_depth}, max_urls={request.max_urls})"
        )
        
        # Discover URLs with requested depth
        discovered_urls = await call_maybe_async(
            scraper_service.discover_url,
            str(request.base_url),
            request.max_depth,
            request.max_urls
        )
        
        if not discovered_urls:
            return {
                "status": "no_urls_found",
                "message": "No URLs discovered from the base URL",
                "base_url": str(request.base_url)
            }
        # Apply domain filter if specified
        if request.domain_filter:
            filtered_urls = [
                url for url in discovered_urls
                if request.domain_filter.lower() in urllib.parse.urlparse(url).netloc.lower()
            ]
            logger.info(
                f"🔍 Filtered {len(discovered_urls)} → {len(filtered_urls)} URLs "
                f"by domain: {request.domain_filter}"
            )
            discovered_urls = filtered_urls
        
        # Filter external links if not allowed
        if not request.follow_external_links:
            base_domain = urllib.parse.urlparse(str(request.base_url)).netloc
            same_domain_urls = [
                url for url in discovered_urls
                if urllib.parse.urlparse(url).netloc == base_domain
            ]
            logger.info(
                f"🔒 Filtered {len(discovered_urls)} → {len(same_domain_urls)} URLs "
                f"(same domain only: {base_domain})"
            )
            discovered_urls = same_domain_urls

        # Merge seed URLs (guarantee critical docs like zones are scraped)
        if request.seed_urls:
            seed_set = {u.strip() for u in request.seed_urls if u and u.strip()}
            before = len(discovered_urls)
            discovered_urls = list(set(discovered_urls) | seed_set)
            logger.info(
                f"🌱 Merged {len(seed_set)} seed URLs: {before} → {len(discovered_urls)} total"
            )

        # Start bulk scraping task
        background_tasks.add_task(
            enhanced_bulk_scrape_task,
            discovered_urls,
            request.auto_store,
            request.max_depth,
            request.extract_deep_images
        )

        return {
            "status": "started",
            "base_url": str(request.base_url),
            "discovered_urls_count": len(discovered_urls),
            "urls_preview": discovered_urls[:10],
            "auto_store": request.auto_store,
            "domain_filter": request.domain_filter,
            "max_depth": request.max_depth,
            "follow_external_links": request.follow_external_links,
            "extract_deep_images": request.extract_deep_images,
            "estimated_time_minutes": round(len(discovered_urls) * 0.5, 1),
            "estimated_documents": len(discovered_urls) * 2}
    
    except Exception as e:
        logger.exception(f"Widget bulk scrape error: {e}")
        raise HTTPException(status_code=500, detail=f"Bulk scrape error: {str(e)}")

@router.post("/widget/upload-file")
async def widget_upload_file(file: UploadFile = File(...),store_in_knowledge: bool = True,chunk_large_files: bool = True):
    """Enhanced file upload with robust processing and storage"""
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="Filename is required")
        
        filename = os.path.basename(file.filename)
        guessed = mimetypes.guess_type(filename)
        content_type = file.content_type or (guessed[0] if guessed else None)
        content_type = content_type or "application/octet-stream"

        logger.info(f"Processing uploaded file: {filename} ({content_type})")
        content = await file.read()
        
        if not content:
            raise HTTPException(status_code=400, detail="File is empty")

        text: Optional[str] = None
        format_type: Optional[str] = None
        metadata: Dict[str, Any] = {}

        # Process file based on type
        try:
            if content_type.startswith("text") or filename.lower().endswith((".txt", ".md")):
                text = content.decode("utf-8", errors="replace")
                format_type = "text"

            elif "pdf" in content_type or filename.lower().endswith(".pdf"):
                pdf_reader = PdfReader(BytesIO(content))
                pages_text = []
                try:
                    if getattr(pdf_reader, "is_encrypted", False):
                        try:
                            pdf_reader.decrypt("")
                        except Exception:
                            logger.debug("PDF is encrypted or decryption failed.")
                    for page in getattr(pdf_reader, "pages", []):
                        page_text = page.extract_text() or ""
                        if page_text.strip():
                            pages_text.append(page_text)
                except Exception as ex_pg:
                    logger.debug(f"PDF page extraction partial failure: {ex_pg}")
                    pages_text = []
                text = "\n\n--- Page Break ---\n\n".join(pages_text)
                format_type = "pdf"
                metadata["total_pages"] = len(getattr(pdf_reader, "pages", []))
                metadata["pages_with_text"] = len(pages_text)

            elif "csv" in content_type or filename.lower().endswith(".csv"):
                csv_text = content.decode("utf-8", errors="replace")
                reader = csv.reader(StringIO(csv_text))
                rows = list(reader)
                if rows:
                    headers = rows[0]
                    data_rows = rows[1:] if len(rows) > 1 else []
                    text_parts = [
                        f"CSV Headers: {', '.join(headers)}",
                        f"Total Rows: {len(data_rows)}",
                        "Sample Data:"
                    ]
                    for i, row in enumerate(data_rows[:10]):
                        text_parts.append(f"Row {i+1}: {', '.join(str(cell) for cell in row)}")
                    text = "\n".join(text_parts)
                else:
                    text = csv_text
                    headers = []
                format_type = "csv"
                metadata["total_rows"] = len(rows)
                metadata["columns"] = len(headers) if rows else 0

            elif ("wordprocessingml" in content_type) or filename.lower().endswith(".docx"):
                doc = Document(BytesIO(content))
                paragraphs = [p.text for p in getattr(doc, "paragraphs", []) if p.text and p.text.strip()]
                text = "\n\n".join(paragraphs)
                format_type = "docx"
                metadata["total_paragraphs"] = len(getattr(doc, "paragraphs", []))
                metadata["paragraphs_with_text"] = len(paragraphs)

            elif ("spreadsheetml" in content_type) or filename.lower().endswith(".xlsx"):
                wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
                sheets_text = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_data = []
                    for row in sheet.iter_rows(values_only=True):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        if any(cell.strip() for cell in row_data):
                            sheet_data.append(", ".join(row_data))
                    if sheet_data:
                        sheets_text.append(f"Sheet: {sheet_name}\n" + "\n".join(sheet_data))
                text = "\n\n--- Sheet Break ---\n\n".join(sheets_text)
                format_type = "xlsx"
                metadata["total_sheets"] = len(getattr(wb, "sheetnames", []))
                metadata["sheets_with_data"] = len(sheets_text)

            elif content_type in ("text/html", "application/xhtml+xml") or filename.lower().endswith((".html", ".htm")):
                soup = BeautifulSoup(content, "html.parser")
                for element in soup(["script", "style"]):
                    element.decompose()
                text = soup.get_text(separator="\n")
                format_type = "html"

            else:
                try:
                    text = content.decode("utf-8", errors="replace")
                    format_type = "unknown"
                except Exception:
                    raise HTTPException(
                        status_code=415,
                        detail=f"Unsupported file type: {content_type}"
                    )

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error processing file {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

        if not text or len(text.strip()) < 10:
            raise HTTPException(status_code=400, detail="File content is too short or unreadable")
        # Normalize whitespace
        text = re.sub(r"\r\n?", "\n", text)
        text = re.sub(r"\n\s*\n", "\n\n", text)
        text = re.sub(r"[ \t]+", " ", text)

        try:
            doc_summary = await call_maybe_async(
                ai_service.generate_summary,
                text,
                max_sentences=4,
                max_chars=800
            )
        except Exception as e:
            logger.debug(f"Summary generation failed: {e}")
            doc_summary = (text[:800] + "...") if len(text) > 800 else text

        response: Dict[str, Any] = {
            "filename": filename,
            "format": format_type,
            "content_length": len(text),
            "word_count": len(text.split()),
            "file_size_bytes": len(content),
            "stored_in_knowledge": False,
            "documents_stored": 0,
            "chunks_created": 0,
            "summary": doc_summary,
            "metadata": metadata,
        }

        async def _store_documents_safe(docs: List[Dict[str, Any]]):
            candidates = ["add_documents", "store_documents", "add_docs", "store", "add"]
            func = None
            for name in candidates:
                if hasattr(postgres_service, name):
                    func = getattr(postgres_service, name)
                    break
            if not func:
                raise RuntimeError("postgres_service has no storage method")

            res = func(docs)
            if inspect.isawaitable(res):
                return await res
            return res

        if store_in_knowledge:
            try:
                if chunk_large_files and len(text) > 2000:
                    chunks = chunk_text(text, chunk_size=1500, overlap=200)
                    documents_to_store = []
                    for i, chunk in enumerate(chunks):
                        documents_to_store.append({
                            "content": chunk,
                            "url": f"file://{filename}#chunk-{i}",
                            "title": f"{filename} (Part {i+1})",
                            "format": format_type,
                            "timestamp": datetime.now().isoformat(),
                            "source": "widget_upload",
                            "images": [],
                            "metadata": metadata,
                        })
                    response["chunks_created"] = len(chunks)
                else:
                    documents_to_store = [{
                        "content": text,
                        "url": f"file://{filename}",
                        "title": filename,
                        "format": format_type,
                        "timestamp": datetime.now().isoformat(),
                        "source": "widget_upload",
                        "images": [],
                        "metadata": metadata,
                    }]
                    response["chunks_created"] = 1

                stored_ids = await _store_documents_safe(documents_to_store)

                if isinstance(stored_ids, (list, tuple, set)):
                    response["documents_stored"] = len(stored_ids)
                    response["stored_in_knowledge"] = len(stored_ids) > 0
                elif isinstance(stored_ids, int):
                    response["documents_stored"] = stored_ids
                    response["stored_in_knowledge"] = stored_ids > 0
                elif stored_ids is None:
                    response["documents_stored"] = 0
                    response["stored_in_knowledge"] = False
                else:
                    try:
                        response["documents_stored"] = len(stored_ids)
                        response["stored_in_knowledge"] = len(stored_ids) > 0
                    except Exception:
                        response["documents_stored"] = 0
                        response["stored_in_knowledge"] = False

            except Exception as e:
                logger.error(f"Error storing file in knowledge base: {e}")
                response["storage_error"] = str(e)

        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Widget upload error: {e}")
        raise HTTPException(status_code=500, detail=f"Upload error: {str(e)}")


@router.get("/widget/knowledge-stats")
async def widget_knowledge_stats():
    """Enhanced knowledge base statistics"""
    try:
        stats = await call_maybe_async(postgres_service.get_collection_stats)
        health = await ai_service.get_service_health()
        services = await orchestration_service.check_service_availability()

        return {
            "document_count": stats.get("document_count", 0) if isinstance(stats, dict) else 0,
            "collection_status": stats.get("status", "unknown") if isinstance(stats, dict) else "unknown",
            "collection_name": stats.get("collection_name", "unknown") if isinstance(stats, dict) else "unknown",
            "search_config": stats.get("search_config", {}) if isinstance(stats, dict) else {},
            "database": stats.get("database", "postgres") if isinstance(stats, dict) else "postgres",
            "connection": stats.get("connection", {}) if isinstance(stats, dict) else {},
            "indexes": stats.get("indexes", []) if isinstance(stats, dict) else [],
            "embedding_dimension": stats.get("embedding_dimension", 0) if isinstance(stats, dict) else 0,
            "ai_services": health.get("service", {}) if isinstance(health, dict) else {},
            "overall_health": health.get("overall_status", "unknown") if isinstance(health, dict) else "unknown",
            "services_availability": services,
            "task_history_count": len(orchestration_service.task_history),
            "last_updated": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.exception(f"Widget stats error: {e}")
        raise HTTPException(status_code=500, detail=f"Stats error: {str(e)}")


@router.delete("/widget/clear-knowledge")
async def widget_clear_knowledge():
    """Clear knowledge base with confirmation"""
    try:
        await call_maybe_async(postgres_service.delete_collection)
        await call_maybe_async(postgres_service.initialize)
        return {
            "status": "success",
            "message": "Knowledge base cleared and reinitialized successfully",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.exception(f"Widget clear error: {e}")
        raise HTTPException(status_code=500, detail=f"Clear error: {str(e)}")


# ===================== Background Tasks =====================

async def store_document_task(docs: List[Dict[str, Any]]):
    """
    Background task to store documents with images.
    
    CRITICAL: Ensures images_json is properly formatted before storage.
    """
    try:
        # Validate and normalize images in each document
        for doc in docs:
            images = doc.get("images", [])
            if not isinstance(images, list):
                doc["images"] = []
                continue
            # Normalize each image
            normalized_images = []
            for img in images:
                if isinstance(img, dict):
                    if img.get("url"):
                        normalized_images.append({
                            "url": img.get("url"),
                            "alt": img.get("alt", ""),
                            "caption": img.get("caption", ""),
                            "type": img.get("type", "content"),
                            "text": img.get("text", "")[:800] 
                        })
                elif isinstance(img, str) and img.startswith("http"):
                    # Convert string URL to dict
                    normalized_images.append({
                        "url": img,
                        "alt": "",
                        "caption": "",
                        "type": "content",
                        "text": ""
                    })
            doc["images"] = normalized_images
            logger.debug(
                f"Normalized {len(normalized_images)} images for document "
                f"{doc.get('url', 'unknown')[:60]}")
        # Store documents
        func = getattr(postgres_service, "add_documents", None)
        if not func:
            raise RuntimeError("postgres_service has no add_documents method")
        result = func(docs)
        if inspect.isawaitable(result):
            await result
        total_images = sum(len(doc.get("images", [])) for doc in docs)
        logger.info(
            f"✅ Stored {len(docs)} documents with {total_images} images total"
        )
    except Exception as e:
        logger.error(f"❌ Failed storing docs: {e}")
        logger.exception("Full traceback:")

async def enhanced_bulk_scrape_task(urls: List[str], auto_store: bool, max_depth: int,extract_images: bool = True):
    """
    Enhanced bulk scraping with improved image extraction and batching.
    Improvements:
    - Adaptive batch sizing based on URL count
    - Better error handling and retry logic
    - Enhanced image extraction and storage
    - Progress logging every 10%
    """
    scraped_count = 0
    stored_count = 0
    error_count = 0
    total_images = 0
    
    if len(urls) > 100:
        batch_size = 10
    elif len(urls) > 50:
        batch_size = 7
    else:
        batch_size = 5

    logger.info(
        f"🚀 Starting DEEP bulk scrape of {len(urls)} URLs "
        f"(depth={max_depth}, batch_size={batch_size}, extract_images={extract_images})"
    )

    for i in range(0, len(urls), batch_size):
        batch = urls[i: i + batch_size]
        batch_num = (i // batch_size) + 1
        total_batches = ((len(urls) - 1) // batch_size) + 1
        progress_pct = (i / len(urls)) * 100
        
        batch_start_time = datetime.now()

        scrape_params = {
            "extract_text": True,
            "extract_links": True, 
            "extract_images": extract_images,
            "extract_tables": True,
            "scroll_page": True,
            "wait_for_element": "body",
            "output_format": "json",
        }
        tasks = [scraper_service.scrape_url(url, scrape_params) for url in batch]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        documents_to_store = []
        batch_errors = []
        batch_images = 0

        for url, result in zip(batch, results):
            if isinstance(result, Exception):
                error_count += 1
                batch_errors.append(f"{url[:50]}: {str(result)[:50]}")
                logger.warning(f"⚠️ Error scraping {url}: {str(result)}")
                continue

            if result.get("status") == "success" and result.get("content"):
                scraped_count += 1
                content = result["content"]
                page_text = content.get("text", "")
                page_images = content.get("images", []) or []
                page_links = content.get("links", []) or []
                
                batch_images += len(page_images)
                total_images += len(page_images)
                if page_text and len(page_text.strip()) >= 100:
                    if auto_store:
                        if len(page_text) > 2500:
                            chunks = chunk_text(page_text, chunk_size=2000, overlap=300)
                            for j, chunk in enumerate(chunks):
                                # First chunk gets all images
                                chunk_images = page_images if j == 0 else []
                                
                                documents_to_store.append({
                                    "content": chunk,
                                    "url": f"{url}#chunk-{j}",
                                    "title": content.get("title", "") or f"Content from {url}",
                                    "format": "text/html",
                                    "timestamp": datetime.now().isoformat(),
                                    "source": "widget_bulk_scrape_deep",
                                    "images": chunk_images,
                                    "metadata": {
                                        "chunk_index": j,
                                        "total_chunks": len(chunks),
                                        "page_images": len(page_images),
                                        "page_links": len(page_links),
                                        "scrape_depth": max_depth
                                    }
                                })
                        else:
                            documents_to_store.append({
                                "content": page_text,
                                "url": url,
                                "title": content.get("title", "") or f"Content from {url}",
                                "format": "text/html",
                                "timestamp": datetime.now().isoformat(),
                                "source": "widget_bulk_scrape_deep",
                                "images": page_images,
                                "metadata": {
                                    "page_images": len(page_images),
                                    "page_links": len(page_links),
                                    "scrape_depth": max_depth
                                }
                            })
                    
                    logger.debug(
                        f"✅ Scraped {url[:60]}: {len(page_text)} chars, "
                        f"{len(page_images)} images, {len(page_links)} links"
                    )
                else:
                    logger.warning(
                        f"⚠️ Skipping {url[:60]}: content too short ({len(page_text)} chars)"
                    )
            else:
                error_count += 1
                logger.warning(f"⚠️ Failed to scrape {url}: {result.get('error', 'Unknown error')}")

        if documents_to_store:
            try:
                store_fn = getattr(postgres_service, "add_documents", None) or getattr(postgres_service, "store_documents", None)
                if not store_fn:
                    raise RuntimeError("No storage function found on postgres_service")
                res = store_fn(documents_to_store)
                if inspect.isawaitable(res):
                    stored_ids = await res
                else:
                    stored_ids = res
                stored_count += len(stored_ids) if stored_ids else 0
                logger.info(f"✅ Batch {i//batch_size + 1}: Stored {len(stored_ids) if stored_ids else 0} documents")
            except Exception as e:
                logger.exception(f"❌ Error storing batch documents: {e}")

        batch_duration = (datetime.now() - batch_start_time).total_seconds()
        logger.info(f"⏱️ Batch {i//batch_size + 1}/{(len(urls)-1)//batch_size + 1} completed in {batch_duration:.1f}s")

        if len(urls) <= 50 or progress_pct % 10 < (batch_size / len(urls)) * 100:
            logger.info(
                f"📊 Progress: {scraped_count}/{len(urls)} scraped, "
                f"{stored_count} stored, {total_images} images, "
                f"{error_count} errors | Batch took {batch_duration:.1f}s")
        if batch_errors:
            logger.warning(f"⚠️ Batch errors: {', '.join(batch_errors[:3])}")
        # Adaptive delay based on batch size
        await asyncio.sleep(min(3.0, batch_size * 0.4))
    # Final summary
    success_rate = (scraped_count / len(urls)) * 100 if urls else 0
    logger.info(
        f"🎉 DEEP bulk scrape completed!\n"
        f"  📄 Scraped: {scraped_count}/{len(urls)} pages ({success_rate:.1f}% success)\n"
        f"  💾 Stored: {stored_count} documents\n"
        f"  🖼️ Images: {total_images} total\n"
        f"  ❌ Errors: {error_count}\n"
        f"  📏 Depth: {max_depth} levels"
    )
